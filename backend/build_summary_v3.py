"""Build v3 review summary — adds Assessment Guidance column to Industry Profiles
and a new Controls Library tab. R-001 to R-014 are fully populated; remaining
themes show placeholder rows pending authoring.

CLI usage:
  Master library (us, all industries, generic scale):
      python build_summary_v3.py
  Tailored user workbook (one industry, NPAT-calibrated scale):
      python build_summary_v3.py --industry FINS --npat 5000000 --out /path/to/out.xlsx
"""
import sys, argparse, os
sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), 'risk-library'))

# ── Tailoring parameters ─────────────────────────────────────────
_parser = argparse.ArgumentParser()
_parser.add_argument("--industry", default=None,
                     help="Industry code (PROF, FINS, TECH, HCAR, RETL, INDU, PUBL). "
                          "If omitted, master library is built (all 7 industries).")
_parser.add_argument("--npat", type=float, default=None,
                     help="Net profit after tax in dollars. If provided, "
                          "Rating Methodology impact scale is calibrated as % of NPAT.")
_parser.add_argument("--out", default="/home/claude/Library_Review_Summary.xlsx",
                     help="Output path for the workbook.")
_parser.add_argument("--config", default=None,
                     help="Path to JSON config file. When provided, all other flags are "
                          "overridden by config values. Used by the FastAPI backend.")
_args, _unknown = _parser.parse_known_args()

# Load config from JSON if provided; otherwise build from individual flags
import json
if _args.config:
    with open(_args.config) as _cf:
        _config_data = json.load(_cf)
else:
    _config_data = {}

# Merge: JSON config takes precedence, then CLI flags, then defaults
_org = _config_data.get("organisation", {}) if _config_data else {}
_fin = _config_data.get("financial", {}) if _config_data else {}

INDUSTRY_FILTER = _config_data.get("industry_code") or _args.industry
NPAT = _fin.get("npat") if _config_data else _args.npat
OUTPUT_PATH = _config_data.get("output_path") or _args.out

# Optional config — only used when --config is provided
ORG_NAME    = _org.get("name", "")
ORG_SIZE    = _org.get("size", "")
ORG_COUNTRY = _org.get("country", "")
ORG_CYCLE   = _org.get("cycle", "")
ORG_NOTES   = _org.get("notes", "")

CURRENCY    = _fin.get("currency", "USD")
FIN_PCTS    = _fin.get("fin_pcts")  # list of 5 percentages, or None for default

FOCUS_AREAS    = _config_data.get("focus_areas", []) if _config_data else []
SCORING_METHOD = _config_data.get("scoring_method", "max") if _config_data else "max"

# RAG thresholds — 4-tier configurable
# Default: Low (1-3), Moderate (4-9), High (10-15), Very High (16-25)
_default_rag = [
    {"label": "Low",       "min": 1,  "max": 3},
    {"label": "Moderate",  "min": 4,  "max": 9},
    {"label": "High",      "min": 10, "max": 15},
    {"label": "Very High", "min": 16, "max": 25},
]
RAG_THRESHOLDS = _config_data.get("rag_thresholds", _default_rag) if _config_data else _default_rag

TAILORED = INDUSTRY_FILTER is not None

# Standard NPAT-based impact bands — defaults
DEFAULT_FIN_PCTS = [0.01, 0.05, 0.15, 0.30, 1.00]  # cumulative thresholds
NPAT_BANDS = [
    (1, "Insignificant",  0.00, FIN_PCTS[0] if FIN_PCTS else DEFAULT_FIN_PCTS[0]),
    (2, "Minor",          FIN_PCTS[0] if FIN_PCTS else DEFAULT_FIN_PCTS[0],
                          FIN_PCTS[1] if FIN_PCTS else DEFAULT_FIN_PCTS[1]),
    (3, "Moderate",       FIN_PCTS[1] if FIN_PCTS else DEFAULT_FIN_PCTS[1],
                          FIN_PCTS[2] if FIN_PCTS else DEFAULT_FIN_PCTS[2]),
    (4, "Major",          FIN_PCTS[2] if FIN_PCTS else DEFAULT_FIN_PCTS[2],
                          FIN_PCTS[3] if FIN_PCTS else DEFAULT_FIN_PCTS[3]),
    (5, "Catastrophic",   FIN_PCTS[3] if FIN_PCTS else DEFAULT_FIN_PCTS[3], None),
]

def _fmt_money(amount):
    """Format dollar amount as $1.5M / $250K / $50K, etc."""
    if amount is None: return ""
    prefix = f"{CURRENCY} " if CURRENCY and CURRENCY != "USD" else "$"
    if amount >= 1_000_000:
        return f"{prefix}{amount/1_000_000:.1f}M".replace(".0M", "M")
    if amount >= 1_000:
        return f"${amount/1_000:.0f}K"
    return f"${amount:.0f}"

# ── Imports of library content ───────────────────────────────────
from risk_library import RISK_LIBRARY, INDUSTRY_NAMES
from risk_controls import CONTROLS, GUIDANCE
# Merge in all the batch additions
import importlib.util
def _load(name, path):
    spec = importlib.util.spec_from_file_location(name, path)
    m = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(m)
    return m
_b5a = _load("b5a", "/home/claude/risk-library/_b5a.py")
_b5b = _load("b5b", "/home/claude/risk-library/_b5b.py")
_g3  = _load("g3",  "/home/claude/risk-library/_g3.py")
_g4  = _load("g4",  "/home/claude/risk-library/_g4.py")
_b6  = _load("b6",  "/home/claude/risk-library/_b6.py")
_b7  = _load("b7",  "/home/claude/risk-library/_b7.py")
_b8a = _load("b8a", "/home/claude/risk-library/_b8a.py")
_b8b = _load("b8b", "/home/claude/risk-library/_b8b.py")
_b9  = _load("b9",  "/home/claude/risk-library/_b9.py")
_b10 = _load("b10", "/home/claude/risk-library/_b10_ai.py")
_fw  = _load("fw",  "/home/claude/risk-library/_frameworks.py")
_th  = _load("th",  "/home/claude/risk-library/_themes.py")
_rn  = _load("rn",  "/home/claude/risk-library/_renames.py")
_dsc = _load("dsc", "/home/claude/risk-library/_descriptions.py")
CONTROLS = {**CONTROLS, **_b5a.CONTROLS_5A, **_b5b.CONTROLS_5B}
GUIDANCE = {**GUIDANCE, **_g3.GUIDANCE_3, **_g4.GUIDANCE_4}

# Merge gap-fill controls into universal controls of relevant themes
for _tid, _new_controls in _b6.GAP_FILL_CONTROLS.items():
    if _tid in CONTROLS:
        CONTROLS[_tid]["universal"] = CONTROLS[_tid]["universal"] + _new_controls

# Merge missing ops controls into universal controls of relevant themes
for _tid, _new_controls in _b7.OPS_CONTROLS.items():
    if _tid in CONTROLS:
        CONTROLS[_tid]["universal"] = CONTROLS[_tid]["universal"] + _new_controls

# Merge completeness gap-fill controls (Batches 8A and 8B) — the hybrid A+B review
for _tid, _new_controls in _b8a.GAP_FILL_8A.items():
    if _tid in CONTROLS:
        CONTROLS[_tid]["universal"] = CONTROLS[_tid]["universal"] + _new_controls
for _tid, _new_controls in _b8b.GAP_FILL_8B.items():
    if _tid in CONTROLS:
        CONTROLS[_tid]["universal"] = CONTROLS[_tid]["universal"] + _new_controls

# Final polish gap-fills (Batch 9)
for _tid, _new_controls in _b9.POLISH_CONTROLS.items():
    if _tid in CONTROLS:
        CONTROLS[_tid]["universal"] = CONTROLS[_tid]["universal"] + _new_controls

# AI driver batch — treat AI as an amplifier of cyber, data, conduct, third-party,
# fraud, compliance, and reputation risk (Option B). R-040 wording unchanged.
for _tid, _new_controls in _b10.AI_DRIVER_CONTROLS.items():
    if _tid in CONTROLS:
        CONTROLS[_tid]["universal"] = CONTROLS[_tid]["universal"] + _new_controls

# Apply control renames — frameworks/policies/standards reframed as the management
# and monitoring of those documents (the actual controls)
_renamed_count = _rn.apply_renames(CONTROLS)
print(f"Renamed {_renamed_count} controls to lead with the management activity.")

# Apply framework enrichment to ALL controls (universal + industry)
def _enrich_control(ctl):
    new_fw = _fw.enrich_framework(ctl["name"], ctl["description"], ctl.get("framework", ""))
    if new_fw and new_fw != ctl.get("framework"):
        ctl = {**ctl, "framework": new_fw}
    return ctl

for _tid, theme in CONTROLS.items():
    theme["universal"] = [_enrich_control(c) for c in theme["universal"]]
    for code in theme["industry"]:
        theme["industry"][code] = [_enrich_control(c) for c in theme["industry"][code]]

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule

wb = Workbook()

# ── Palette ──
C_NAVY  = "0F2A47"; C_MBLUE = "3B7BC4"
C_LBLUE = "D6E4F0"; C_LBLUE2= "EBF5FB"
C_WHITE = "FFFFFF"; C_LGREY = "F2F2F2"; C_DGREY = "7F7F7F"
C_PALE_NAVY = "E8EEF5"

SCORE_BG = {1:"D5F5E3", 2:"EAF7E1", 3:"FEF9E7", 4:"FDEBD0", 5:"FADBD8"}

# 4-tier risk rating methodology
# Low (1-3) | Moderate (4-9) | High (10-15) | Very High (16-25)
TIER_BG = {
    "Low":       "D5F5E3",   # green
    "Moderate":  "FDEBD0",   # amber/orange
    "High":      "E74C3C",   # red
    "Very High": "7B241C",   # dark red / maroon
}
TIER_FG = {
    "Low":       "196F3D",
    "Moderate":  "7E5109",
    "High":      "FFFFFF",   # white text on red
    "Very High": "FFFFFF",   # white text on maroon
}

# Legacy 3-tier kept only for any leftover references; new code uses tier_for()
RAG_BG = {"GREEN":"D5F5E3", "AMBER":"FDEBD0", "RED":"FADBD8"}
RAG_FG = {"GREEN":"196F3D", "AMBER":"7E5109", "RED":"7B1F14"}

# 5-point label scales
IMPACT_LABELS     = {1:"Insignificant",  2:"Minor",      3:"Moderate", 4:"Major",  5:"Catastrophic"}
LIKELIHOOD_LABELS = {1:"Rare",           2:"Unlikely",   3:"Possible", 4:"Likely", 5:"Almost Certain"}

# Dropdown option strings — used in data validation
IMPACT_OPTIONS     = [f"{IMPACT_LABELS[i]} ({i})" for i in range(1, 6)]
LIKELIHOOD_OPTIONS = [f"{LIKELIHOOD_LABELS[i]} ({i})" for i in range(1, 6)]

CONTROL_TYPE_BG = {
    "Preventive":"D6E4F0", "Detective":"FEF9E7",
    "Corrective":"FDEBD0", "Directive":"E8EEF5"
}
CONTROL_NATURE_BG = {
    "Manual":"FFFFFF", "Automated":"D5F5E3", "Semi-automated":"FEF9E7"
}

INDUSTRY_ORDER = ["PROF", "FINS", "TECH", "HCAR", "RETL", "INDU", "PUBL"]

# Validate industry code if tailored
if TAILORED and INDUSTRY_FILTER not in INDUSTRY_ORDER:
    raise SystemExit(f"Invalid --industry '{INDUSTRY_FILTER}'. Must be one of: {', '.join(INDUSTRY_ORDER)}")

# Filtered list — single industry when tailored, all 7 when master
ACTIVE_INDUSTRIES = [INDUSTRY_FILTER] if TAILORED else INDUSTRY_ORDER

def bdr():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)
def fill(c): return PatternFill("solid", fgColor=c)
def hf(sz=11, bold=True, col=C_WHITE): return Font(name="Arial", size=sz, bold=bold, color=col)
def cf(sz=10, bold=False, col="000000"): return Font(name="Arial", size=sz, bold=bold, color=col)
def ctr(): return Alignment(horizontal="center", vertical="center", wrap_text=True)
def lft(): return Alignment(horizontal="left", vertical="top", wrap_text=True)

def tier_for(score):
    """4-tier risk score classification using configurable thresholds."""
    for tier in RAG_THRESHOLDS:
        if tier["min"] <= score <= tier["max"]:
            return tier["label"]
    # Fallback: return the highest tier label if score exceeds all maxes
    return RAG_THRESHOLDS[-1]["label"]

def rag_for(score):
    """Legacy 3-tier — used by hidden matrix sheets only."""
    if score >= 15: return "RED"
    if score >= 6:  return "AMBER"
    return "GREEN"

def build_causes(v):
    return (f"PEOPLE: {v['causes_people']}\n"
            f"PROCESS: {v['causes_process']}\n"
            f"SYSTEMS: {v['causes_systems']}\n"
            f"EXTERNAL: {v['causes_external']}")

def build_impacts(v, theme_domain, theme_id):
    rationale = v["impact_rationale"]
    consequence = v["consequence"]
    channel_map = {
        "Financial":   ["Direct financial loss", "Cash flow / earnings impact"],
        "Customer":    ["Customer detriment / dissatisfaction", "Customer attrition"],
        "Regulatory":  ["Regulatory action / fines / enforcement", "Licence or accreditation impact"],
        "Reputational":["Brand and reputation damage", "Loss of stakeholder trust"],
        "People":      ["Staff harm or wellbeing impact", "Workforce attrition or capability loss"],
        "Social":      ["Community / stakeholder detriment", "Social licence impact"],
        "Environmental":["Environmental harm or compliance breach", "Long-term ecological impact"],
    }
    parts = [f"FINANCIAL: {consequence}"]
    primary = rationale.split("—")[0].strip().split(" ")[0]
    if primary in channel_map:
        ch = channel_map[primary]
        parts.append(f"PRIMARY ({primary.upper()}): {ch[0]}; {ch[1]}.")

    opp_cost = {
        "R-001":"Lost market share, declining win rates, foregone revenue from competitive position erosion",
        "R-002":"Foregone revenue from disruption-driven displacement; cost of late catch-up investment",
        "R-003":"Foregone synergies or growth from failed deal; capital tied up in underperforming asset",
        "R-005":"Foregone diversified revenue; concentration premium loss",
        "R-008":"Lost transactions, billing leakage, customer revenue not captured",
        "R-010":"Foregone revenue from declined work or unfulfilled demand",
        "R-011":"Lost repeat business; foregone referrals from quality issues",
        "R-013":"Foregone growth and revenue from capability gaps",
    }
    if theme_id in opp_cost:
        parts.append(f"OPPORTUNITY COST: {opp_cost[theme_id]}.")
    if theme_domain not in ["Financial", "Geopolitical and Economic"]:
        parts.append("OPERATIONAL: Service / process disruption; remediation effort and management distraction.")
    legal = {
        "R-006":"Civil / criminal proceedings against perpetrators; possible director liability",
        "R-011":"PI / negligence claims; product liability litigation",
        "R-014":"Discrimination / harassment claims; class actions",
    }
    if theme_id in legal:
        parts.append(f"LEGAL / LITIGATION: {legal[theme_id]}.")
    parts.append(f"STAKEHOLDER: {rationale}.")
    return "\n".join(parts)

def build_guidance(theme_id, archetype):
    """Return assessment guidance text. Returns placeholder if not yet authored."""
    if theme_id in GUIDANCE and archetype in GUIDANCE[theme_id]:
        g = GUIDANCE[theme_id][archetype]
        return (f"DATA: {g['data']}\n"
                f"SYSTEMS: {g['systems']}\n"
                f"STAKEHOLDERS: {g['stakeholders']}\n"
                f"KEY FACTORS: {g['key_factors']}")
    return "(Detailed guidance pending — review approach on R-001 to R-014 for format.)"

# ════════════════════════════════════════════════════════════════
# SHEET 1 – OVERVIEW
# ════════════════════════════════════════════════════════════════
ws_cover = wb.active
ws_cover.sheet_view.showGridLines = False
ws_cover.column_dimensions["A"].width = 3
ws_cover.column_dimensions["B"].width = 32
ws_cover.column_dimensions["C"].width = 70
ws_cover.column_dimensions["D"].width = 12
ws_cover.column_dimensions["E"].width = 12

if TAILORED:
    # ── PERSONALISED LANDING PAGE ─────────────────────────────────
    ws_cover.title = "Welcome"

    # Hero — personalised with org name when available
    ws_cover.merge_cells("B2:E3")
    c = ws_cover["B2"]
    if ORG_NAME:
        c.value = f"Risk Register  |  {ORG_NAME}"
    else:
        c.value = "Welcome to your tailored Blue Peak Risk Register"
    c.font = Font(name="Arial", size=18, bold=True, color=C_WHITE)
    c.fill = fill(C_NAVY); c.alignment = ctr()
    ws_cover.row_dimensions[2].height = 22; ws_cover.row_dimensions[3].height = 22

    ws_cover.merge_cells("B4:E4")
    c = ws_cover["B4"]
    # Build subtitle from available org details
    subtitle_parts = [INDUSTRY_NAMES[INDUSTRY_FILTER]]
    if ORG_SIZE: subtitle_parts.append(ORG_SIZE)
    if ORG_COUNTRY: subtitle_parts.append(ORG_COUNTRY)
    if NPAT: subtitle_parts.append(f"NPAT {_fmt_money(NPAT)}")
    c.value = "  ·  ".join(subtitle_parts)
    c.font = Font(name="Arial", size=11, italic=True, color=C_WHITE)
    c.fill = fill(C_MBLUE); c.alignment = ctr()
    ws_cover.row_dimensions[4].height = 18

    # Section: This is your tailored register
    ws_cover.merge_cells("B6:E6")
    c = ws_cover["B6"]
    c.value = "WHAT WE'VE TAILORED FOR YOU"
    c.font = hf(sz=12, col=C_NAVY); c.fill = fill(C_LBLUE); c.alignment = lft()
    ws_cover.row_dimensions[6].height = 20

    tailored_rows = [
        ("Industry",
         f"{INDUSTRY_NAMES[INDUSTRY_FILTER]} — every risk in this workbook is described, classified, and controlled for your sector."),
        ("Impact scale",
         (f"Calibrated to your NPAT of {_fmt_money(NPAT)}. The Rating Methodology & Heatmap tab "
          f"shows the dollar thresholds for each impact level, computed as a percentage of NPAT.")
         if NPAT else
         "Generic financial language — you can refine the dollar thresholds yourself in the Rating Methodology & Heatmap tab."),
        ("Rating methodology",
         f"4-tier inherent risk rating using your configured thresholds: " +
         ", ".join(f"{t['label']} ({t['min']}-{t['max']})" for t in RAG_THRESHOLDS) +
         f". Inherent Impact aggregated using {SCORING_METHOD.upper()} of 7 categories."),
        ("Risk coverage",
         "All 40 risks across 14 domains — strategic, financial, operational, people, technology, compliance, "
         "reputational, environmental, safety, project, third-party, customer, geopolitical, and financial crime."),
        ("Controls",
         "Each risk comes with a recommended control set tailored to your industry, classified by theme, "
         "type, nature, and frequency, with framework citations (NIST, ISO, COSO, sector regulations)."),
        ("Assessment-ready",
         "Drop-down menus for likelihood and 7 individual impact categories. The Inherent Impact and Risk "
         "Rating columns calculate automatically using your selections."),
    ]
    if ORG_CYCLE:
        tailored_rows.append(("Review cycle", f"You set a {ORG_CYCLE.lower()} review cycle for this register."))
    if FOCUS_AREAS:
        tailored_rows.append(("Focus areas", f"You flagged: {', '.join(FOCUS_AREAS)}. All 40 risks are included; "
                                              "use these as priorities for review."))
    for i, (label, val) in enumerate(tailored_rows, 7):
        bg = C_LGREY if i % 2 == 0 else C_WHITE
        c1 = ws_cover.cell(row=i, column=2, value=label)
        c1.font = cf(10, bold=True); c1.fill = fill(bg); c1.alignment = lft(); c1.border = bdr()
        c2 = ws_cover.cell(row=i, column=3, value=val)
        c2.font = cf(10); c2.fill = fill(bg); c2.alignment = lft(); c2.border = bdr()
        ws_cover.merge_cells(f"C{i}:E{i}")
        ws_cover.row_dimensions[i].height = 36

    # Section: Quick Start
    qs_start = 7 + len(tailored_rows) + 1
    ws_cover.merge_cells(f"B{qs_start}:E{qs_start}")
    c = ws_cover.cell(row=qs_start, column=2, value="QUICK START — 3 STEPS")
    c.font = hf(sz=12, col=C_NAVY); c.fill = fill(C_LBLUE); c.alignment = lft()
    ws_cover.row_dimensions[qs_start].height = 20

    qs_steps = [
        ("1.  Review your risk profile",
         "Open Complete Risk Profile. Each row is a risk tailored to your industry. "
         "Read the Description, Causes, Impacts, and Assessment Guidance to understand what's in scope."),
        ("2.  Rate inherent risk",
         "For each risk, select Likelihood from the dropdown (column K), and choose impact ratings for each "
         "of the 7 impact categories (columns L–R). The workbook calculates Inherent Impact (max of 7) "
         "and Inherent Risk Rating automatically."),
        ("3.  Assess your controls",
         "In Control Environment, review each suggested control. Set Design Effectiveness and "
         "(if Design is Effective) Operating Effectiveness from the dropdowns. Control Effectiveness "
         "is calculated for you. Filter by risk, theme, or owner to focus your review."),
    ]
    for i, (step, val) in enumerate(qs_steps, qs_start + 1):
        bg = C_LGREY if i % 2 == 0 else C_WHITE
        c1 = ws_cover.cell(row=i, column=2, value=step)
        c1.font = cf(10, bold=True); c1.fill = fill(bg); c1.alignment = lft(); c1.border = bdr()
        c2 = ws_cover.cell(row=i, column=3, value=val)
        c2.font = cf(10); c2.fill = fill(bg); c2.alignment = lft(); c2.border = bdr()
        ws_cover.merge_cells(f"C{i}:E{i}")
        ws_cover.row_dimensions[i].height = 44

    # Section: Tabs in this workbook
    tabs_start = qs_start + len(qs_steps) + 2
    ws_cover.merge_cells(f"B{tabs_start}:E{tabs_start}")
    c = ws_cover.cell(row=tabs_start, column=2, value="TABS IN THIS WORKBOOK")
    c.font = hf(sz=12, col=C_NAVY); c.fill = fill(C_LBLUE); c.alignment = lft()
    ws_cover.row_dimensions[tabs_start].height = 20

    sheets_info = [
        ("Welcome",                       "This page — your guide to the workbook."),
        ("Risk Hierarchy",                "The universal risk taxonomy — Level 1 (domain) and Level 2 (theme) — with descriptions and default treatments."),
        ("Complete Risk Profile",         "Your main assessment tab. 40 risks tailored to your industry, with assessment guidance and dropdown-driven ratings."),
        ("Rating Methodology & Heatmap",  "The 4-tier rating scheme with calibrated impact scale, the 5×5 reference grid, and a live heatmap of your inherent risks."),
        ("Control Environment",           "Recommended controls per risk with effectiveness rating workflow (Design → Operating → Control)."),
        ("Controls Library",              "All controls grouped by functional theme (e.g. Monitoring, Limits, Investigation) for cross-risk view."),
    ]
    for i, (sheet, purpose) in enumerate(sheets_info, tabs_start + 1):
        bg = C_LGREY if i % 2 == 0 else C_WHITE
        c1 = ws_cover.cell(row=i, column=2, value=sheet)
        c1.font = cf(10, bold=True); c1.fill = fill(bg); c1.alignment = lft(); c1.border = bdr()
        c2 = ws_cover.cell(row=i, column=3, value=purpose)
        c2.font = cf(10); c2.fill = fill(bg); c2.alignment = lft(); c2.border = bdr()
        ws_cover.merge_cells(f"C{i}:E{i}")
        ws_cover.row_dimensions[i].height = 30

    # Section: Adding more
    addmore_start = tabs_start + len(sheets_info) + 2
    ws_cover.merge_cells(f"B{addmore_start}:E{addmore_start}")
    c = ws_cover.cell(row=addmore_start, column=2, value="EXTENDING THIS REGISTER")
    c.font = hf(sz=12, col=C_NAVY); c.fill = fill(C_LBLUE); c.alignment = lft()
    ws_cover.row_dimensions[addmore_start].height = 20

    addmore_text = (
        "This register is a starting point — not the final word. You can add risks specific to your "
        "organisation by inserting rows in Complete Risk Profile (use any unused Risk ID format such as R-101+). "
        "You can add additional controls in Control Environment using your own Control IDs. "
        "Hidden tabs (used by the Blue Peak Risk team for reference) can be unhidden via right-click on any tab "
        "if you'd like to see the underlying matrices and frameworks library."
    )
    c = ws_cover.cell(row=addmore_start + 1, column=2, value="Add what you need")
    c.font = cf(10, bold=True); c.fill = fill(C_LGREY); c.alignment = lft(); c.border = bdr()
    c2 = ws_cover.cell(row=addmore_start + 1, column=3, value=addmore_text)
    c2.font = cf(10); c2.fill = fill(C_LGREY); c2.alignment = lft(); c2.border = bdr()
    ws_cover.merge_cells(f"C{addmore_start + 1}:E{addmore_start + 1}")
    ws_cover.row_dimensions[addmore_start + 1].height = 60

else:
    # ── MASTER LIBRARY OVERVIEW (for us) ──────────────────────────
    ws_cover.title = "Overview"

    ws_cover.merge_cells("B2:E3")
    c = ws_cover["B2"]
    c.value = "Blue Peak Risk — Library Review (v3)"
    c.font = Font(name="Arial", size=18, bold=True, color=C_WHITE)
    c.fill = fill(C_NAVY); c.alignment = ctr()
    ws_cover.row_dimensions[2].height = 22; ws_cover.row_dimensions[3].height = 22

    ws_cover.merge_cells("B4:E4")
    c = ws_cover["B4"]
    c.value = "40 themes  ·  14 domains  ·  7 industry archetypes  ·  280 variants  ·  Assessment guidance + Controls library"
    c.font = Font(name="Arial", size=11, italic=True, color=C_WHITE)
    c.fill = fill(C_MBLUE); c.alignment = ctr()
    ws_cover.row_dimensions[4].height = 18

    ws_cover.merge_cells("B6:E6")
    c = ws_cover["B6"]
    c.value = "WHAT'S IN THIS WORKBOOK"
    c.font = hf(sz=12, col=C_NAVY); c.fill = fill(C_LBLUE); c.alignment = lft()
    ws_cover.row_dimensions[6].height = 20

    sheets_info = [
        ("1. Overview",         "This sheet — workbook guide and library statistics"),
        ("2. Risk Hierarchy",   "Level 1 and Level 2 risk types — the universal taxonomy with descriptions and default treatments"),
        ("3. Complete Risk Profile","Full risk content per industry — descriptions, causes, impacts, assessment guidance, ratings, owner"),
        ("4. Rating Methodology & Heatmap","4-tier risk rating scheme with definitions, 5×5 reference grid, and live 40×7 heatmap"),
        ("5. Control Environment", "Suggested controls per risk × industry with theme, type, nature, KCIs, framework, and effectiveness ratings"),
        ("6. Controls Library", "Functional categorisation of controls — control themes with consolidated control names"),
    ]
    for i, (sheet, purpose) in enumerate(sheets_info, 7):
        bg = C_LGREY if i % 2 == 0 else C_WHITE
        c1 = ws_cover.cell(row=i, column=2, value=sheet)
        c1.font = cf(10, bold=True); c1.fill = fill(bg); c1.alignment = lft(); c1.border = bdr()
        c2 = ws_cover.cell(row=i, column=3, value=purpose)
        c2.font = cf(10); c2.fill = fill(bg); c2.alignment = lft(); c2.border = bdr()
        ws_cover.merge_cells(f"C{i}:E{i}")
        ws_cover.row_dimensions[i].height = 18

    # Authoring status
    ws_cover.merge_cells("B16:E16")
    c = ws_cover["B16"]
    c.value = "AUTHORING STATUS"
    c.font = hf(sz=12, col=C_NAVY); c.fill = fill(C_LBLUE); c.alignment = lft()
    ws_cover.row_dimensions[16].height = 20

    themes_with_controls = sorted(CONTROLS.keys())

    status_rows = [
        ("Themes with full controls + guidance", f"{len(themes_with_controls)} of 40 — COMPLETE"),
        ("Total industry-tailored controls (incl. universal)", f"{sum(len(c['universal']) + sum(len(v) for v in c['industry'].values()) for c in CONTROLS.values())} control entries"),
        ("Total assessment guidance entries", f"{sum(len(g) for g in GUIDANCE.values())} entries (40 themes × 7 industries)"),
        ("Domains covered", f"14 (incl. Financial Crime — ORX-aligned)"),
        ("Industries", "7 archetypes (Professional, Financial, Tech/SaaS, Healthcare, Retail, Industrial, Public/NFP)"),
    ]
    for i, (label, val) in enumerate(status_rows, 17):
        bg = C_LGREY if i % 2 == 0 else C_WHITE
        c1 = ws_cover.cell(row=i, column=2, value=label)
        c1.font = cf(10, bold=True); c1.fill = fill(bg); c1.alignment = lft(); c1.border = bdr()
        c2 = ws_cover.cell(row=i, column=3, value=val)
        c2.font = cf(10); c2.fill = fill(bg); c2.alignment = lft(); c2.border = bdr()
        ws_cover.merge_cells(f"C{i}:E{i}")
        ws_cover.row_dimensions[i].height = 24

# ════════════════════════════════════════════════════════════════
# SHEETS 2-5 — Themes Index, Likelihood Matrix, Impact Matrix, Risk Score Matrix
# (same as before)
# ════════════════════════════════════════════════════════════════
ws_idx = wb.create_sheet("Risk Hierarchy")
ws_idx.sheet_view.showGridLines = False
ws_idx.freeze_panes = "A3"
for col, w in [("A",8),("B",24),("C",55),("D",70),("E",18),("F",18)]:
    ws_idx.column_dimensions[col].width = w
ws_idx.merge_cells("A1:F1")
c = ws_idx["A1"]
c.value = "Risk Hierarchy — Level 1 and Level 2 Risk Types"
c.font = Font(name="Arial", size=14, bold=True, color=C_WHITE)
c.fill = fill(C_NAVY); c.alignment = ctr()
ws_idx.row_dimensions[1].height = 24
hdrs = ["Risk ID", "Risk Level 1", "Risk Level 2", "Description", "Default Risk Treatment", "Primary Impact"]
for j, h in enumerate(hdrs, 1):
    c = ws_idx.cell(row=2, column=j, value=h)
    c.font = hf(sz=10); c.fill = fill(C_MBLUE); c.alignment = ctr(); c.border = bdr()
ws_idx.row_dimensions[2].height = 32
for i, r in enumerate(RISK_LIBRARY, 3):
    bg = C_LGREY if i % 2 == 0 else C_WHITE
    prof = r["industries"]["PROF"]
    # Replace 'Transfer' with 'Share' per ISO 31000:2018
    treatment = r["treatment"].replace("Transfer", "Share")
    description = _dsc.UNIVERSAL_DESCRIPTIONS.get(r["id"], "")
    vals = [r["id"], r["domain"], r["theme"], description, treatment, prof["impact_rationale"]]
    for j, val in enumerate(vals, 1):
        c = ws_idx.cell(row=i, column=j, value=val)
        c.font = cf(9, bold=(j==1)); c.fill = fill(bg)
        c.alignment = ctr() if j in [1,5] else lft(); c.border = bdr()
    ws_idx.row_dimensions[i].height = 44

def build_rating_matrix(ws, title, key):
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "C3"
    for col, w in [("A",7),("B",46)] + [(get_column_letter(i),12) for i in range(3,10)]:
        ws.column_dimensions[col].width = w
    ws.merge_cells("A1:I1")
    c = ws["A1"]; c.value = title
    c.font = Font(name="Arial", size=14, bold=True, color=C_WHITE)
    c.fill = fill(C_NAVY); c.alignment = ctr()
    ws.row_dimensions[1].height = 24
    hdrs = ["Risk ID", "Universal Theme"] + [INDUSTRY_NAMES[code] for code in INDUSTRY_ORDER]
    for j, h in enumerate(hdrs, 1):
        c = ws.cell(row=2, column=j, value=h)
        c.font = hf(sz=9); c.fill = fill(C_MBLUE); c.alignment = ctr(); c.border = bdr()
    ws.row_dimensions[2].height = 38
    for i, r in enumerate(RISK_LIBRARY, 3):
        bg = C_LGREY if i % 2 == 0 else C_WHITE
        ws.cell(row=i, column=1, value=r["id"]).font = cf(9, bold=True)
        ws.cell(row=i, column=1).fill = fill(bg); ws.cell(row=i, column=1).alignment = ctr(); ws.cell(row=i, column=1).border = bdr()
        ws.cell(row=i, column=2, value=r["theme"]).font = cf(9)
        ws.cell(row=i, column=2).fill = fill(bg); ws.cell(row=i, column=2).alignment = lft(); ws.cell(row=i, column=2).border = bdr()
        for j, code in enumerate(INDUSTRY_ORDER, 3):
            val = r["industries"][code][key]
            c = ws.cell(row=i, column=j, value=val)
            c.font = cf(11, bold=True); c.fill = fill(SCORE_BG[val])
            c.alignment = ctr(); c.border = bdr()
        ws.row_dimensions[i].height = 24

_ws_lik = wb.create_sheet("Likelihood Matrix")
build_rating_matrix(_ws_lik, "Inherent Likelihood Matrix (1-5)", "inh_likelihood")
_ws_imp = wb.create_sheet("Impact Matrix")
build_rating_matrix(_ws_imp, "Inherent Impact Matrix (1-5)", "inh_impact")

# Risk Score Matrix
ws_s = wb.create_sheet("Risk Score Matrix")
ws_s.sheet_view.showGridLines = False
ws_s.freeze_panes = "C3"
for col, w in [("A",7),("B",46)] + [(get_column_letter(i),12) for i in range(3,10)]:
    ws_s.column_dimensions[col].width = w
ws_s.merge_cells("A1:I1")
c = ws_s["A1"]
c.value = "Inherent Risk Score Matrix (L × I, 1-25)  —  Heat map"
c.font = Font(name="Arial", size=14, bold=True, color=C_WHITE)
c.fill = fill(C_NAVY); c.alignment = ctr()
ws_s.row_dimensions[1].height = 24
hdrs = ["Risk ID", "Universal Theme"] + [INDUSTRY_NAMES[code] for code in INDUSTRY_ORDER]
for j, h in enumerate(hdrs, 1):
    c = ws_s.cell(row=2, column=j, value=h)
    c.font = hf(sz=9); c.fill = fill(C_MBLUE); c.alignment = ctr(); c.border = bdr()
ws_s.row_dimensions[2].height = 38
for i, r in enumerate(RISK_LIBRARY, 3):
    bg = C_LGREY if i % 2 == 0 else C_WHITE
    ws_s.cell(row=i, column=1, value=r["id"]).font = cf(9, bold=True)
    ws_s.cell(row=i, column=1).fill = fill(bg); ws_s.cell(row=i, column=1).alignment = ctr(); ws_s.cell(row=i, column=1).border = bdr()
    ws_s.cell(row=i, column=2, value=r["theme"]).font = cf(9)
    ws_s.cell(row=i, column=2).fill = fill(bg); ws_s.cell(row=i, column=2).alignment = lft(); ws_s.cell(row=i, column=2).border = bdr()
    for j, code in enumerate(INDUSTRY_ORDER, 3):
        ind = r["industries"][code]
        score = ind["inh_likelihood"] * ind["inh_impact"]
        rag = rag_for(score)
        c = ws_s.cell(row=i, column=j, value=score)
        c.font = Font(name="Arial", size=11, bold=True, color=RAG_FG[rag])
        c.fill = fill(RAG_BG[rag]); c.alignment = ctr(); c.border = bdr()
    ws_s.row_dimensions[i].height = 24

# ════════════════════════════════════════════════════════════════
# SHEET 6 — INDUSTRY PROFILES (with new Assessment Guidance column)
# Columns: ID | Domain | Title | Description | Causes | Impacts | ASSESSMENT GUIDANCE | InhL | InhI | Score | RAG | Owner
# ════════════════════════════════════════════════════════════════
ws_p = wb.create_sheet("Complete Risk Profile")
ws_p.sheet_view.showGridLines = False
ws_p.freeze_panes = "C3"

# Column layout: 21 columns
# A Risk ID | B Industry | C Risk Title | D Risk Level 1 | E Risk Level 2 |
# F Description | G Suggested Owner | H Causes | I Impacts | J Assessment Guidance |
# K Inh L | L-R Inh Impact ratings (7 categories) | S Inh Overall I | T Score | U RAG
cols = [
    ("A",7),    # Risk ID
    ("B",22),   # Industry
    ("C",36),   # Risk Title (industry-tailored)
    ("D",18),   # Risk Type Level 1
    ("E",36),   # Risk Type Level 2
    ("F",42),   # Description (industry-tailored)
    ("G",22),   # Suggested Owner (moved up)
    ("H",48),   # Causes
    ("I",46),   # Impacts
    ("J",50),   # Assessment Guidance
    ("K",18),   # Likelihood (dropdown)
    ("L",18),   # Inh — Financial
    ("M",18),   # Inh — Customer
    ("N",18),   # Inh — Regulatory
    ("O",18),   # Inh — Reputational
    ("P",18),   # Inh — People
    ("Q",18),   # Inh — Social
    ("R",18),   # Inh — Environmental
    ("S",18),   # Inherent Impact (formula)
    ("T",22),   # Inherent Risk Rating (formula)
]
for col, w in cols:
    ws_p.column_dimensions[col].width = w

ws_p.merge_cells("A1:T1")
c = ws_p["A1"]
c.value = "Complete Risk Profile — full risk content per industry (description, causes, impacts, assessment guidance, ratings, owner)"
c.font = Font(name="Arial", size=14, bold=True, color=C_WHITE)
c.fill = fill(C_NAVY); c.alignment = ctr()
ws_p.row_dimensions[1].height = 24

hdrs = [
    "Risk ID","Industry","Risk Title","Risk Type Level 1","Risk Type Level 2",
    "Description","Suggested Owner",
    "Causes (People/Process/Systems/External)","Impacts","Assessment Guidance",
    "Likelihood",
    "Inh — Financial","Inh — Customer","Inh — Regulatory","Inh — Reputational",
    "Inh — People","Inh — Social","Inh — Environmental",
    "Inherent Impact","Inherent Risk Rating"
]
for j, h in enumerate(hdrs, 1):
    c = ws_p.cell(row=2, column=j, value=h)
    c.font = hf(sz=9); c.fill = fill(C_MBLUE); c.alignment = ctr(); c.border = bdr()
ws_p.row_dimensions[2].height = 44

# Map "primary impact" string in library to which of the 7 columns it should populate
PRIMARY_IMPACT_COL = {
    "Financial": 12,      # L
    "Customer": 13,       # M
    "Regulatory": 14,     # N
    "Reputational": 15,   # O
    "People": 16,         # P
    "Social": 17,         # Q
    "Environmental": 18,  # R
}

current_row = 3
# Per request: industry as a column, no separator rows. Iterate by industry × risk.
for code in ACTIVE_INDUSTRIES:
    for r in RISK_LIBRARY:
        ind = r["industries"][code]
        bg = C_LGREY if current_row % 2 == 0 else C_WHITE

        causes = build_causes(ind)
        impacts = build_impacts(ind, r["domain"], r["id"])
        guidance = build_guidance(r["id"], code)

        # Determine which of 7 impact-rating columns to seed from existing primary impact
        primary = ind.get("impact_rationale", "")
        primary_key = next((k for k in PRIMARY_IMPACT_COL if primary.startswith(k)), None)
        primary_col = PRIMARY_IMPACT_COL.get(primary_key)

        # Pre-populate Likelihood with the library's value (user can change)
        likelihood_str = f"{LIKELIHOOD_LABELS[ind['inh_likelihood']]} ({ind['inh_likelihood']})"

        # Pre-populate ONLY the primary impact column with the library value; other 6 blank
        def impact_str_or_blank(col_index):
            if col_index == primary_col:
                return f"{IMPACT_LABELS[ind['inh_impact']]} ({ind['inh_impact']})"
            return None

        # Formula for Inherent Impact (col S):
        # Extract numbers from L:R (cells like "Major (4)") via MID(...,FIND("(",...)+1,...) approach,
        # then aggregate per scoring_method:
        #   "max" / "weighted" -> MAX of the 7 (highest single impact dominates — risk-mgmt convention)
        #   "avg"              -> AVERAGE of non-blank impacts (rounded to nearest integer)
        # Then map to IMPACT_LABELS.
        rownum = current_row
        # Helper expression that parses one cell into a number, or 0 if blank
        def parse_cell_num(cell_ref):
            return (f'IFERROR(VALUE(MID({cell_ref},FIND("(",{cell_ref})+1,'
                    f'FIND(")",{cell_ref})-FIND("(",{cell_ref})-1)),0)')

        impact_cells = [parse_cell_num(f"{c}{rownum}") for c in ("L","M","N","O","P","Q","R")]

        if SCORING_METHOD == "avg":
            # Sum of non-zero values divided by count of non-zero values, rounded
            sum_expr = "+".join(impact_cells)
            count_expr = "+".join([f'IF({e}>0,1,0)' for e in impact_cells])
            agg_impact_num = f'IF(({count_expr})=0,0,ROUND(({sum_expr})/({count_expr}),0))'
        else:  # "max" or "weighted" both fall through to MAX
            agg_impact_num = f'MAX({",".join(impact_cells)})'

        # Map number to impact label
        impact_label_map = (
            f'IF({agg_impact_num}=0,"",'
            f'IF({agg_impact_num}=1,"Insignificant (1)",'
            f'IF({agg_impact_num}=2,"Minor (2)",'
            f'IF({agg_impact_num}=3,"Moderate (3)",'
            f'IF({agg_impact_num}=4,"Major (4)",'
            f'IF({agg_impact_num}=5,"Catastrophic (5)",""))))))'
        )
        inh_impact_formula = f'={impact_label_map}'

        # Formula for Inherent Risk Rating (col T):
        # Extract Likelihood number from K, multiply by aggregated impact number,
        # then map to tier label using CONFIGURABLE RAG_THRESHOLDS.
        likelihood_num = parse_cell_num(f"K{rownum}")
        score_expr = f'({likelihood_num})*({agg_impact_num})'

        # Build nested IF from highest tier down (since IF chains short-circuit)
        # Sort tiers by min descending, so highest threshold is evaluated first
        _sorted_tiers = sorted(RAG_THRESHOLDS, key=lambda t: t["min"], reverse=True)
        # Construct: IF(score>=tier1.min, "tier1.label (score)", IF(score>=tier2.min, ..., ...))
        rating_chain = f'"{_sorted_tiers[-1]["label"]} ("&{score_expr}&")"'  # innermost (lowest tier)
        for tier in _sorted_tiers[:-1][::-1]:
            # Each iteration wraps the chain in another IF (we go from lowest to next-up)
            rating_chain = f'IF({score_expr}>={tier["min"]},"{tier["label"]} ("&{score_expr}&")",{rating_chain})'
        # Outermost: blank if either factor is zero
        risk_rating_formula = (
            f'=IF(OR({likelihood_num}=0,{agg_impact_num}=0),"",{rating_chain})'
        )

        # Build the 20-column row
        cells = [
            (1, r["id"],                  cf(9, bold=True), ctr()),
            (2, INDUSTRY_NAMES[code],     cf(9, bold=True), lft()),
            (3, ind["title"],             cf(9, bold=True), lft()),
            (4, r["domain"],              cf(9), lft()),
            (5, r["theme"],               cf(9), lft()),
            (6, ind["description"],       cf(9), lft()),
            (7, ind["owner_suggested"],   cf(9), lft()),
            (8, causes,                   cf(9), lft()),
            (9, impacts,                  cf(9), lft()),
            (10, guidance,                cf(9), lft()),
            (11, likelihood_str,          cf(10, bold=True), ctr()),  # K
            (12, impact_str_or_blank(12), cf(10, bold=True), ctr()),  # L Financial
            (13, impact_str_or_blank(13), cf(10, bold=True), ctr()),  # M Customer
            (14, impact_str_or_blank(14), cf(10, bold=True), ctr()),  # N Regulatory
            (15, impact_str_or_blank(15), cf(10, bold=True), ctr()),  # O Reputational
            (16, impact_str_or_blank(16), cf(10, bold=True), ctr()),  # P People
            (17, impact_str_or_blank(17), cf(10, bold=True), ctr()),  # Q Social
            (18, impact_str_or_blank(18), cf(10, bold=True), ctr()),  # R Environmental
            (19, inh_impact_formula,      cf(10, bold=True), ctr()),  # S formula
            (20, risk_rating_formula,     cf(10, bold=True), ctr()),  # T formula
        ]
        for col, val, font_, align in cells:
            c = ws_p.cell(row=current_row, column=col, value=val)
            c.font = font_; c.alignment = align; c.border = bdr()
            if col == 10:
                c.fill = fill(C_PALE_NAVY) if r["id"] in GUIDANCE else fill(C_LGREY)
            else:
                c.fill = fill(bg)
        ws_p.row_dimensions[current_row].height = 130
        current_row += 1

# Enable autofilter on Complete Risk Profile (now 20 columns)
ws_p.auto_filter.ref = f"A2:T{current_row - 1}"

# ── Data validation: dropdowns on Likelihood (K) and 7 impact columns (L:R) ──
likelihood_options_str = '"' + ",".join(LIKELIHOOD_OPTIONS) + '"'
impact_options_str = '"' + ",".join(IMPACT_OPTIONS) + '"'

dv_likelihood = DataValidation(type="list", formula1=likelihood_options_str, allow_blank=True)
dv_likelihood.error = "Please select from the dropdown"
dv_likelihood.errorTitle = "Invalid entry"
dv_likelihood.prompt = "Select likelihood: Rare (1) → Almost Certain (5)"
dv_likelihood.promptTitle = "Likelihood"
ws_p.add_data_validation(dv_likelihood)
dv_likelihood.add(f"K3:K{current_row - 1}")

for col_letter in ("L","M","N","O","P","Q","R"):
    dv_imp = DataValidation(type="list", formula1=impact_options_str, allow_blank=True)
    dv_imp.error = "Please select from the dropdown"
    dv_imp.errorTitle = "Invalid entry"
    dv_imp.prompt = "Select impact: Insignificant (1) → Catastrophic (5). Leave blank if not material."
    dv_imp.promptTitle = "Impact rating"
    ws_p.add_data_validation(dv_imp)
    dv_imp.add(f"{col_letter}3:{col_letter}{current_row - 1}")

# ── Conditional formatting: colour Likelihood + 7 impact + Inherent Impact cells by trailing number ──
# Score 1 → Low (green), 2 → green tint, 3 → amber, 4 → red, 5 → maroon
fill_imp_1 = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")
fill_imp_2 = PatternFill(start_color="EAF7E1", end_color="EAF7E1", fill_type="solid")
fill_imp_3 = PatternFill(start_color="FDEBD0", end_color="FDEBD0", fill_type="solid")
fill_imp_4 = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")
fill_imp_5 = PatternFill(start_color="7B241C", end_color="7B241C", fill_type="solid")

# For each rating column (K through S), colour based on trailing "(n)" digit in the value
rating_cols = ["K","L","M","N","O","P","Q","R","S"]
for col_letter in rating_cols:
    rng = f"{col_letter}3:{col_letter}{current_row - 1}"
    ws_p.conditional_formatting.add(rng,
        FormulaRule(formula=[f'ISNUMBER(SEARCH("(1)",{col_letter}3))'], fill=fill_imp_1))
    ws_p.conditional_formatting.add(rng,
        FormulaRule(formula=[f'ISNUMBER(SEARCH("(2)",{col_letter}3))'], fill=fill_imp_2))
    ws_p.conditional_formatting.add(rng,
        FormulaRule(formula=[f'ISNUMBER(SEARCH("(3)",{col_letter}3))'], fill=fill_imp_3))
    ws_p.conditional_formatting.add(rng,
        FormulaRule(formula=[f'ISNUMBER(SEARCH("(4)",{col_letter}3))'], fill=fill_imp_4))
    ws_p.conditional_formatting.add(rng,
        FormulaRule(formula=[f'ISNUMBER(SEARCH("(5)",{col_letter}3))'], fill=fill_imp_5))

# Inherent Risk Rating (col T) — colour by tier label
fill_low  = PatternFill(start_color=TIER_BG["Low"], end_color=TIER_BG["Low"], fill_type="solid")
fill_mod  = PatternFill(start_color=TIER_BG["Moderate"], end_color=TIER_BG["Moderate"], fill_type="solid")
fill_high = PatternFill(start_color=TIER_BG["High"], end_color=TIER_BG["High"], fill_type="solid")
fill_vh   = PatternFill(start_color=TIER_BG["Very High"], end_color=TIER_BG["Very High"], fill_type="solid")

rng_t = f"T3:T{current_row - 1}"
ws_p.conditional_formatting.add(rng_t,
    FormulaRule(formula=[f'ISNUMBER(SEARCH("Low",T3))'], fill=fill_low))
ws_p.conditional_formatting.add(rng_t,
    FormulaRule(formula=[f'ISNUMBER(SEARCH("Moderate",T3))'], fill=fill_mod))
ws_p.conditional_formatting.add(rng_t,
    FormulaRule(formula=[f'ISNUMBER(SEARCH("Very High",T3))'], fill=fill_vh))
# "High" matches both "High" and "Very High" by SEARCH — apply Very High first then High
ws_p.conditional_formatting.add(rng_t,
    FormulaRule(formula=[f'AND(ISNUMBER(SEARCH("High",T3)),NOT(ISNUMBER(SEARCH("Very High",T3))))'], fill=fill_high))


# ════════════════════════════════════════════════════════════════
# RATING METHODOLOGY & HEATMAP — defines the 4-tier scheme and shows live state
# ════════════════════════════════════════════════════════════════
ws_m = wb.create_sheet("Rating Methodology & Heatmap")
ws_m.sheet_view.showGridLines = False

# Column widths
for col, w in [("A",6),("B",22),("C",55),("D",10),("E",10),("F",10),("G",10),("H",10),("I",18)]:
    ws_m.column_dimensions[col].width = w

# Title
ws_m.merge_cells("A1:I1")
c = ws_m["A1"]
c.value = "Rating Methodology & Heatmap — 4-tier risk rating scheme with live 5×5 reference and 40×7 view"
c.font = Font(name="Arial", size=14, bold=True, color=C_WHITE)
c.fill = fill(C_NAVY); c.alignment = ctr()
ws_m.row_dimensions[1].height = 24

# ── SECTION 1: Likelihood Scale ──
m_row = 3
ws_m.merge_cells(f"A{m_row}:I{m_row}")
c = ws_m.cell(row=m_row, column=1, value="1.  Likelihood Scale")
c.font = Font(name="Arial", size=12, bold=True, color=C_WHITE)
c.fill = fill(C_MBLUE); c.alignment = lft()
ws_m.row_dimensions[m_row].height = 22
m_row += 1

likelihood_defs = [
    (1, "Rare",            "Could occur only in exceptional circumstances; not expected in the foreseeable future. Indicative frequency: less than once in 10 years."),
    (2, "Unlikely",        "Could occur at some time but not expected. Indicative frequency: once in 5–10 years."),
    (3, "Possible",        "Might occur at some time. Indicative frequency: once in 2–5 years."),
    (4, "Likely",          "Will probably occur in most circumstances. Indicative frequency: at least once in 2 years."),
    (5, "Almost Certain",  "Expected to occur in most circumstances; recurring. Indicative frequency: more than once a year."),
]
hdrs_l = ["Score", "Likelihood", "Definition"]
for j, h in enumerate(hdrs_l, 1):
    c = ws_m.cell(row=m_row, column=j, value=h)
    c.font = hf(sz=10); c.fill = fill(C_NAVY); c.alignment = ctr(); c.border = bdr()
ws_m.row_dimensions[m_row].height = 22
m_row += 1
for n, label, defn in likelihood_defs:
    bg = C_LGREY if m_row % 2 == 0 else C_WHITE
    cells = [(1, n, cf(10, bold=True), ctr()),
             (2, label, cf(10, bold=True), ctr()),
             (3, defn, cf(9), lft())]
    ws_m.merge_cells(start_row=m_row, start_column=3, end_row=m_row, end_column=9)
    for col, val, font_, align in cells:
        c = ws_m.cell(row=m_row, column=col, value=val)
        c.font = font_; c.alignment = align; c.border = bdr(); c.fill = fill(bg)
    ws_m.row_dimensions[m_row].height = 32
    m_row += 1

m_row += 1  # spacer

# ── SECTION 2: Impact Scale ──
ws_m.merge_cells(f"A{m_row}:I{m_row}")
c = ws_m.cell(row=m_row, column=1, value="2.  Impact Scale")
c.font = Font(name="Arial", size=12, bold=True, color=C_WHITE)
c.fill = fill(C_MBLUE); c.alignment = lft()
ws_m.row_dimensions[m_row].height = 22
m_row += 1

if NPAT:
    # Build NPAT-calibrated definitions: each tier shows the dollar threshold
    # plus the qualitative description. Bands are 0-1%, 1-5%, 5-15%, 15-30%, >30% of NPAT.
    def _band_text(score):
        for n, label, lo, hi in NPAT_BANDS:
            if n == score:
                lo_amt = lo * NPAT
                hi_amt = hi * NPAT if hi is not None else None
                if hi_amt is None:
                    return f"Financial loss > {_fmt_money(lo_amt)} (>{int(lo*100)}% NPAT)"
                if lo == 0:
                    return f"Financial loss < {_fmt_money(hi_amt)} (<{int(hi*100)}% NPAT)"
                return (f"Financial loss {_fmt_money(lo_amt)}–{_fmt_money(hi_amt)} "
                        f"({int(lo*100)}–{int(hi*100)}% NPAT)")
        return ""
    impact_defs = [
        (1, "Insignificant",  f"{_band_text(1)}; no customer impact; internal process only; no regulatory consequence."),
        (2, "Minor",          f"{_band_text(2)}; isolated customer complaints; short, contained operational disruption; minor regulator interest."),
        (3, "Moderate",       f"{_band_text(3)}; multiple customers affected; service disruption requiring escalation; reportable to regulator."),
        (4, "Major",          f"{_band_text(4)}; widespread customer harm or remediation; multi-day operational impact; regulatory action; brand damage."),
        (5, "Catastrophic",   f"{_band_text(5)}; threat to viability or continuity; severe customer harm; major regulatory enforcement; loss of licence to operate."),
    ]
else:
    impact_defs = [
        (1, "Insignificant",  "Negligible financial loss; no customer impact; internal process only; no regulatory consequence."),
        (2, "Minor",          "Small financial loss; isolated customer complaints; short, contained operational disruption; minor regulator interest."),
        (3, "Moderate",       "Material financial loss; multiple customers affected; service disruption requiring escalation; reportable to regulator."),
        (4, "Major",          "Significant financial loss; widespread customer harm or remediation; multi-day operational impact; regulatory action; brand damage."),
        (5, "Catastrophic",   "Threat to viability or continuity; severe customer harm; major regulatory enforcement; loss of licence to operate; severe brand damage."),
    ]
hdrs_i = ["Score", "Impact", "Definition"]
for j, h in enumerate(hdrs_i, 1):
    c = ws_m.cell(row=m_row, column=j, value=h)
    c.font = hf(sz=10); c.fill = fill(C_NAVY); c.alignment = ctr(); c.border = bdr()
ws_m.row_dimensions[m_row].height = 22
m_row += 1
for n, label, defn in impact_defs:
    bg = C_LGREY if m_row % 2 == 0 else C_WHITE
    cells = [(1, n, cf(10, bold=True), ctr()),
             (2, label, cf(10, bold=True), ctr()),
             (3, defn, cf(9), lft())]
    ws_m.merge_cells(start_row=m_row, start_column=3, end_row=m_row, end_column=9)
    for col, val, font_, align in cells:
        c = ws_m.cell(row=m_row, column=col, value=val)
        c.font = font_; c.alignment = align; c.border = bdr(); c.fill = fill(bg)
    ws_m.row_dimensions[m_row].height = 32
    m_row += 1

m_row += 1

# ── SECTION 3: Tier Definitions ──
ws_m.merge_cells(f"A{m_row}:I{m_row}")
c = ws_m.cell(row=m_row, column=1, value="3.  Inherent Risk Rating Tiers (Likelihood × Impact)")
c.font = Font(name="Arial", size=12, bold=True, color=C_WHITE)
c.fill = fill(C_MBLUE); c.alignment = lft()
ws_m.row_dimensions[m_row].height = 22
m_row += 1

tier_defs = [
    ("Low",       "1 – 3",    "Acceptable level of risk. Standard controls are sufficient. Monitor periodically."),
    ("Moderate",  "4 – 9",    "Generally acceptable but warrants management attention. Verify control effectiveness."),
    ("High",      "10 – 15",  "Material risk requiring active management. Strengthen controls; report to executive."),
    ("Very High", "16 – 25",  "Unacceptable level of risk. Immediate executive / board attention; aggressive treatment required."),
]
hdrs_t = ["", "Tier", "Score Range", "Description"]
for j, h in enumerate(hdrs_t, 1):
    c = ws_m.cell(row=m_row, column=j, value=h)
    c.font = hf(sz=10); c.fill = fill(C_NAVY); c.alignment = ctr(); c.border = bdr()
ws_m.row_dimensions[m_row].height = 22
m_row += 1
for tier, score_range, defn in tier_defs:
    cells = [(1, "", cf(10), ctr()),
             (2, tier, cf(10, bold=True, col=TIER_FG[tier]), ctr()),
             (3, score_range, cf(10, bold=True), ctr()),
             (4, defn, cf(9), lft())]
    ws_m.merge_cells(start_row=m_row, start_column=4, end_row=m_row, end_column=9)
    for col, val, font_, align in cells:
        c = ws_m.cell(row=m_row, column=col, value=val)
        c.font = font_; c.alignment = align; c.border = bdr()
        if col == 1:
            c.fill = fill(TIER_BG[tier])
        elif col == 2:
            c.fill = fill(TIER_BG[tier])
        else:
            c.fill = fill(C_WHITE)
    ws_m.row_dimensions[m_row].height = 30
    m_row += 1

m_row += 1

# ── SECTION 4: 5x5 Reference Heatmap ──
ws_m.merge_cells(f"A{m_row}:I{m_row}")
c = ws_m.cell(row=m_row, column=1, value="4.  5×5 Risk Score Reference Grid")
c.font = Font(name="Arial", size=12, bold=True, color=C_WHITE)
c.fill = fill(C_MBLUE); c.alignment = lft()
ws_m.row_dimensions[m_row].height = 22
m_row += 1

# Header row: Likelihood labels across the top
ws_m.cell(row=m_row, column=1, value="").fill = fill(C_WHITE)
ws_m.merge_cells(start_row=m_row, start_column=2, end_row=m_row, end_column=2)
c = ws_m.cell(row=m_row, column=2, value="Impact ↓ / Likelihood →")
c.font = hf(sz=9); c.fill = fill(C_NAVY); c.alignment = ctr(); c.border = bdr()
for L in range(1, 6):
    c = ws_m.cell(row=m_row, column=2 + L, value=f"{LIKELIHOOD_LABELS[L]}\n({L})")
    c.font = hf(sz=9); c.fill = fill(C_NAVY); c.alignment = ctr(); c.border = bdr()
ws_m.row_dimensions[m_row].height = 32
m_row += 1

# Render grid: row per impact (5 down to 1), col per likelihood (1 to 5)
for I in range(5, 0, -1):
    c = ws_m.cell(row=m_row, column=1, value="")
    c.fill = fill(C_WHITE); c.border = bdr()
    c = ws_m.cell(row=m_row, column=2, value=f"{IMPACT_LABELS[I]} ({I})")
    c.font = hf(sz=9, col=C_WHITE); c.fill = fill(C_NAVY); c.alignment = ctr(); c.border = bdr()
    for L in range(1, 6):
        score = L * I
        tier = tier_for(score)
        c = ws_m.cell(row=m_row, column=2 + L, value=f"{tier}\n({score})")
        c.font = Font(name="Arial", size=10, bold=True, color=TIER_FG[tier])
        c.fill = fill(TIER_BG[tier]); c.alignment = ctr(); c.border = bdr()
    ws_m.row_dimensions[m_row].height = 36
    m_row += 1

m_row += 1

# ── SECTION 5: Live 40 × 7 heatmap ──
ws_m.merge_cells(f"A{m_row}:I{m_row}")
c = ws_m.cell(row=m_row, column=1, value="5.  Live Inherent Risk Heatmap — 40 risks × 7 industries (auto-updates from Complete Risk Profile)")
c.font = Font(name="Arial", size=12, bold=True, color=C_WHITE)
c.fill = fill(C_MBLUE); c.alignment = lft()
ws_m.row_dimensions[m_row].height = 22
m_row += 1

# Adjust col widths for heatmap section: col A skinny, B = risk label, C-I = 7 industries
ws_m.column_dimensions["B"].width = 38

# Header
ws_m.cell(row=m_row, column=1, value="").fill = fill(C_WHITE)
c = ws_m.cell(row=m_row, column=2, value="Risk")
c.font = hf(sz=9); c.fill = fill(C_NAVY); c.alignment = ctr(); c.border = bdr()
for j, code in enumerate(ACTIVE_INDUSTRIES, 3):
    c = ws_m.cell(row=m_row, column=j, value=code)
    c.font = hf(sz=9); c.fill = fill(C_NAVY); c.alignment = ctr(); c.border = bdr()
ws_m.row_dimensions[m_row].height = 24
m_row += 1

# One row per risk (40), one column per active industry — values pulled live from Complete Risk Profile.
# Complete Risk Profile rows iterate as ACTIVE_INDUSTRIES × RISK_LIBRARY: 40 rows per industry block.
profile_start_row = 3
for risk_idx, r in enumerate(RISK_LIBRARY):
    bg = C_LGREY if m_row % 2 == 0 else C_WHITE
    c = ws_m.cell(row=m_row, column=1, value="")
    c.fill = fill(bg); c.border = bdr()
    c = ws_m.cell(row=m_row, column=2, value=f"{r['id']} — {r['theme']}")
    c.font = cf(9, bold=True); c.fill = fill(bg); c.alignment = lft(); c.border = bdr()
    for ind_idx, code in enumerate(ACTIVE_INDUSTRIES):
        cr_row = profile_start_row + ind_idx * len(RISK_LIBRARY) + risk_idx
        formula = f"='Complete Risk Profile'!T{cr_row}"
        col_idx = 3 + ind_idx
        c = ws_m.cell(row=m_row, column=col_idx, value=formula)
        c.font = cf(9, bold=True); c.alignment = ctr(); c.border = bdr()
        c.fill = fill(C_WHITE)
    ws_m.row_dimensions[m_row].height = 24
    m_row += 1

# Conditional formatting on the live heatmap — covers as many cols as ACTIVE_INDUSTRIES
heatmap_first_data_row = m_row - len(RISK_LIBRARY)
heatmap_last_col_letter = get_column_letter(2 + len(ACTIVE_INDUSTRIES))
heatmap_rng = f"C{heatmap_first_data_row}:{heatmap_last_col_letter}{m_row - 1}"
ws_m.conditional_formatting.add(heatmap_rng,
    FormulaRule(formula=[f'ISNUMBER(SEARCH("Low",C{heatmap_first_data_row}))'], fill=fill_low))
ws_m.conditional_formatting.add(heatmap_rng,
    FormulaRule(formula=[f'ISNUMBER(SEARCH("Moderate",C{heatmap_first_data_row}))'], fill=fill_mod))
ws_m.conditional_formatting.add(heatmap_rng,
    FormulaRule(formula=[f'ISNUMBER(SEARCH("Very High",C{heatmap_first_data_row}))'], fill=fill_vh))
ws_m.conditional_formatting.add(heatmap_rng,
    FormulaRule(formula=[f'AND(ISNUMBER(SEARCH("High",C{heatmap_first_data_row})),NOT(ISNUMBER(SEARCH("Very High",C{heatmap_first_data_row}))))'], fill=fill_high))


# ════════════════════════════════════════════════════════════════
# SHEET 7 — CONTROLS LIBRARY (NEW)
# Columns: Risk ID | Industry | Risk Title | Control ID | Name | Type | Nature
#          | Description | Frequency | Owner | KRIs | Framework
# ════════════════════════════════════════════════════════════════
ws_c = wb.create_sheet("Control Environment")
ws_c.sheet_view.showGridLines = False
ws_c.freeze_panes = "A3"

# 17 columns total — description and owner moved up; 3 new effectiveness columns at the end
cols = [
    ("A",7),    # Risk ID
    ("B",26),   # Industry (full name)
    ("C",38),   # Risk title
    ("D",26),   # Control Applicability
    ("E",14),   # Control ID
    ("F",32),   # Control name
    ("G",55),   # Description (moved up)
    ("H",24),   # Suggested Owner (moved up)
    ("I",28),   # Control Theme
    ("J",13),   # Type
    ("K",14),   # Nature
    ("L",18),   # Frequency
    ("M",50),   # Control Monitoring / Key Control Indicators
    ("N",32),   # Framework
    ("O",18),   # Design Effectiveness Rating (NEW)
    ("P",18),   # Operating Effectiveness Rating (NEW)
    ("Q",18),   # Control Effectiveness Rating (NEW — formula)
]
for col, w in cols:
    ws_c.column_dimensions[col].width = w

ws_c.merge_cells("A1:Q1")
c = ws_c["A1"]
c.value = "Control Environment — suggested controls per risk × industry with effectiveness assessment"
c.font = Font(name="Arial", size=14, bold=True, color=C_WHITE)
c.fill = fill(C_NAVY); c.alignment = ctr()
ws_c.row_dimensions[1].height = 24

hdrs = [
    "Risk ID","Industry","Risk Title","Control Applicability","Control ID","Control Name",
    "Description","Suggested Owner",
    "Control Theme","Type","Nature","Frequency",
    "Control Monitoring / Key Control Indicators","Framework",
    "Design Effectiveness Rating","Operating Effectiveness Rating","Control Effectiveness Rating"
]
for j, h in enumerate(hdrs, 1):
    c = ws_c.cell(row=2, column=j, value=h)
    c.font = hf(sz=9); c.fill = fill(C_MBLUE); c.alignment = ctr(); c.border = bdr()
ws_c.row_dimensions[2].height = 36

current_row = 3

for r in RISK_LIBRARY:
    rid = r["id"]
    if rid not in CONTROLS:
        # Placeholder row - one per active industry, indicating pending authoring
        for code in ACTIVE_INDUSTRIES:
            ind = r["industries"][code]
            placeholder_cells = [
                (1, rid, cf(9, bold=True), ctr()),
                (2, INDUSTRY_NAMES[code], cf(9, bold=True), lft()),
                (3, ind["title"], cf(9), lft()),
                (4, "—", cf(9), ctr()),
                (5, "—", cf(9), ctr()),
                (6, "(Controls pending — review R-001 to R-014 for format)", cf(9, col="7F7F7F"), lft()),
                (7, "Detailed controls to be authored in next batch.", cf(9, col="7F7F7F"), lft()),
                (8, "—", cf(9), lft()),
                (9, "—", cf(9), ctr()),
                (10, "—", cf(9), ctr()),
                (11, "—", cf(9), ctr()),
                (12, "—", cf(9), ctr()),
                (13, "—", cf(9), lft()),
                (14, "—", cf(9), ctr()),
                (15, None, cf(9), ctr()),
                (16, None, cf(9), ctr()),
                (17, None, cf(9), ctr()),
            ]
            bg = C_LGREY
            for col, val, font_, align in placeholder_cells:
                c = ws_c.cell(row=current_row, column=col, value=val)
                c.font = font_; c.alignment = align; c.border = bdr()
                c.fill = fill(bg)
            ws_c.row_dimensions[current_row].height = 22
            current_row += 1
        continue

    # Theme has full controls — render universal + industry-specific per active industry
    universal = CONTROLS[rid]["universal"]
    for code in ACTIVE_INDUSTRIES:
        industry_specific = CONTROLS[rid]["industry"][code]
        all_controls = universal + industry_specific
        ind = r["industries"][code]

        for ctl_idx, ctl in enumerate(all_controls, 1):
            ctl_id = f"C-{rid[2:]}-{code}-{ctl_idx:02d}"
            scope = "Universal" if ctl in universal else f"Specific ({INDUSTRY_NAMES[code]})"
            bg = C_LGREY if current_row % 2 == 0 else C_WHITE

            # Control Effectiveness formula:
            #   Design blank             -> "Not Assessed"
            #   Design <> "Effective"    -> Control Effectiveness = Design rating
            #   Design = "Effective" + Operating blank -> "Not Assessed"
            #   Design = "Effective" + Operating set   -> Control Effectiveness = Operating
            ce_formula = (
                f'=IF(O{current_row}="","Not Assessed",'
                f'IF(O{current_row}<>"Effective",O{current_row},'
                f'IF(P{current_row}="","Not Assessed",P{current_row})))'
            )

            cells = [
                (1, rid, cf(9, bold=True), ctr()),
                (2, INDUSTRY_NAMES[code], cf(9, bold=True), lft()),
                (3, ind["title"], cf(9), lft()),
                (4, scope, cf(9, col="3B7BC4" if scope == "Universal" else "7E5109"), ctr()),
                (5, ctl_id, cf(9, bold=True), ctr()),
                (6, ctl["name"], cf(9, bold=True), lft()),
                (7, ctl["description"], cf(9), lft()),
                (8, ctl["owner"], cf(9), lft()),
                (9, _th.classify_theme(ctl["name"], ctl["description"]), cf(9), ctr()),
                (10, ctl["type"], cf(9), ctr()),
                (11, ("Semi-automated" if ctl["nature"] == "Hybrid" else ctl["nature"]), cf(9), ctr()),
                (12, ctl["frequency"], cf(9), ctr()),
                (13, ctl["kris"], cf(9), lft()),
                (14, ctl["framework"] or "—", cf(9), ctr()),
                (15, None, cf(9), ctr()),                  # Design Effectiveness — left blank for user
                (16, None, cf(9), ctr()),                  # Operating Effectiveness — left blank for user
                (17, ce_formula, cf(9, bold=True), ctr()), # Control Effectiveness — formula
            ]
            for col, val, font_, align in cells:
                c = ws_c.cell(row=current_row, column=col, value=val)
                c.font = font_; c.alignment = align; c.border = bdr()
                if col == 10:
                    c.fill = fill(CONTROL_TYPE_BG.get(val, bg))
                elif col == 11:
                    c.fill = fill(CONTROL_NATURE_BG.get(val, bg))
                elif col in (15, 16, 17):
                    # Pale fill on the rating columns to suggest "fillable"
                    c.fill = fill("F8F9FA")
                else:
                    c.fill = fill(bg)
            ws_c.row_dimensions[current_row].height = 60
            current_row += 1

# Data validation: dropdown lists for Design (col O) and Operating (col P) effectiveness
dv_options = '"Effective,Partially Effective,Ineffective"'
dv_design = DataValidation(type="list", formula1=dv_options, allow_blank=True)
dv_design.error = "Please select Effective, Partially Effective, or Ineffective"
dv_design.errorTitle = "Invalid entry"
dv_design.prompt = "Select effectiveness rating"
dv_design.promptTitle = "Design Effectiveness"
ws_c.add_data_validation(dv_design)
dv_design.add(f"O3:O{current_row - 1}")

dv_operating = DataValidation(type="list", formula1=dv_options, allow_blank=True)
dv_operating.error = "Please select Effective, Partially Effective, or Ineffective"
dv_operating.errorTitle = "Invalid entry"
dv_operating.prompt = "Only assess Operating Effectiveness when Design Effectiveness is rated 'Effective'. Otherwise leave blank — Control Effectiveness will inherit the Design rating."
dv_operating.promptTitle = "Operating Effectiveness"
ws_c.add_data_validation(dv_operating)
dv_operating.add(f"P3:P{current_row - 1}")

# Conditional formatting on the three rating columns — colour by rating value
# Effective = green, Partially Effective = amber, Ineffective = red, Not Assessed = grey
fill_eff = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
fill_part = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
fill_ineff = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
fill_na = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")

for col_letter in ("O", "P", "Q"):
    rng = f"{col_letter}3:{col_letter}{current_row - 1}"
    ws_c.conditional_formatting.add(rng,
        FormulaRule(formula=[f'{col_letter}3="Effective"'], fill=fill_eff))
    ws_c.conditional_formatting.add(rng,
        FormulaRule(formula=[f'{col_letter}3="Partially Effective"'], fill=fill_part))
    ws_c.conditional_formatting.add(rng,
        FormulaRule(formula=[f'{col_letter}3="Ineffective"'], fill=fill_ineff))
    ws_c.conditional_formatting.add(rng,
        FormulaRule(formula=[f'{col_letter}3="Not Assessed"'], fill=fill_na))

# Operating column (P): grey out when Design (O) is blank OR not "Effective"
# This visually signals "not applicable — leave blank"
ws_c.conditional_formatting.add(f"P3:P{current_row - 1}",
    FormulaRule(formula=[f'OR(O3="",O3<>"Effective")'], fill=fill_na))

# Enable autofilter on Control Environment header for easy filtering
ws_c.auto_filter.ref = f"A2:Q{current_row - 1}"

# ════════════════════════════════════════════════════════════════
# SHEET 8 — FRAMEWORKS REFERENCE (NEW)
# ════════════════════════════════════════════════════════════════
ws_f = wb.create_sheet("Frameworks Reference")
ws_f.sheet_view.showGridLines = False
ws_f.freeze_panes = "A3"

cols = [
    ("A", 28),  # Category
    ("B", 32),  # Framework
    ("C", 36),  # Section / Reference
    ("D", 80),  # Description
]
for col, w in cols:
    ws_f.column_dimensions[col].width = w

ws_f.merge_cells("A1:D1")
c = ws_f["A1"]
c.value = "Frameworks Reference — standards and frameworks cited in the Controls Library"
c.font = Font(name="Arial", size=14, bold=True, color=C_WHITE)
c.fill = fill(C_NAVY); c.alignment = ctr()
ws_f.row_dimensions[1].height = 24

hdrs_f = ["Category", "Framework", "Section / Reference", "Brief description"]
for j, h in enumerate(hdrs_f, 1):
    c = ws_f.cell(row=2, column=j, value=h)
    c.font = hf(sz=10); c.fill = fill(C_MBLUE); c.alignment = ctr(); c.border = bdr()
ws_f.row_dimensions[2].height = 28

# Group by category for visual separation
current_category = None
fw_row = 3
for cat, fw_name, section, desc in _b7.FRAMEWORKS_REFERENCE:
    bg = C_LGREY if fw_row % 2 == 0 else C_WHITE
    cells = [
        (1, cat, cf(9, bold=(cat != current_category)), lft()),
        (2, fw_name, cf(9, bold=True), lft()),
        (3, section, cf(9), lft()),
        (4, desc, cf(9), lft()),
    ]
    for col, val, font_, align in cells:
        c = ws_f.cell(row=fw_row, column=col, value=val)
        c.font = font_; c.alignment = align; c.border = bdr()
        c.fill = fill(bg)
    ws_f.row_dimensions[fw_row].height = 38
    current_category = cat
    fw_row += 1

# Enable autofilter on Frameworks Reference too
ws_f.auto_filter.ref = f"A2:D{fw_row - 1}"

# Hide the Frameworks Reference sheet — internal/reference only
ws_f.sheet_state = "hidden"

# ════════════════════════════════════════════════════════════════
# SHEET 9 — CONTROL THEMES REFERENCE (NEW)
# ════════════════════════════════════════════════════════════════
ws_t = wb.create_sheet("Controls Library")
ws_t.sheet_view.showGridLines = False
ws_t.freeze_panes = "A3"

cols_t = [
    ("A", 24),  # Group
    ("B", 30),  # Control Theme
    ("C", 50),  # Definition
    ("D", 45),  # Control Name
    ("E", 70),  # Description (NEW)
]
for col, w in cols_t:
    ws_t.column_dimensions[col].width = w

ws_t.merge_cells("A1:E1")
c = ws_t["A1"]
c.value = "Controls Library — functional categorisation of controls in the Control Environment"
c.font = Font(name="Arial", size=14, bold=True, color=C_WHITE)
c.fill = fill(C_NAVY); c.alignment = ctr()
ws_t.row_dimensions[1].height = 24

hdrs_t = ["Group", "Control Theme", "Definition", "Control Name", "Description"]
for j, h in enumerate(hdrs_t, 1):
    c = ws_t.cell(row=2, column=j, value=h)
    c.font = hf(sz=10); c.fill = fill(C_MBLUE); c.alignment = ctr(); c.border = bdr()
ws_t.row_dimensions[2].height = 28

# Build consolidated map: theme -> sorted unique control names (across all 40 risks × 7 industries)
# Also build name -> representative description (longest version found is most informative)
from collections import defaultdict
_theme_to_controls = defaultdict(set)
_name_to_description = {}

def _consider(name, description):
    """Keep the longest description seen for any control name (the most informative)."""
    if name not in _name_to_description or len(description) > len(_name_to_description[name]):
        _name_to_description[name] = description

for _tid, _theme_data in CONTROLS.items():
    for _ctl in _theme_data["universal"]:
        _t = _th.classify_theme(_ctl["name"], _ctl["description"])
        _theme_to_controls[_t].add(_ctl["name"])
        _consider(_ctl["name"], _ctl["description"])
    for _code, _ctl_list in _theme_data["industry"].items():
        for _ctl in _ctl_list:
            _t = _th.classify_theme(_ctl["name"], _ctl["description"])
            _theme_to_controls[_t].add(_ctl["name"])
            _consider(_ctl["name"], _ctl["description"])

t_row = 3
for grp, theme, defn, examples in _th.THEME_DEFINITIONS:
    control_names_set = _theme_to_controls.get(theme, set())
    sorted_names = sorted(control_names_set) if control_names_set else ["(no controls currently classified to this theme)"]

    for ctl_name in sorted_names:
        bg = C_LGREY if t_row % 2 == 0 else C_WHITE
        # Look up the representative description; fall back to empty for the placeholder string
        ctl_description = _name_to_description.get(ctl_name, "")
        # Populate Group, Theme, Definition on EVERY row so filtering works cleanly
        cells = [
            (1, grp, cf(9, bold=True), lft()),
            (2, theme, cf(9, bold=True, col="0F2A47"), lft()),
            (3, defn, cf(9), lft()),
            (4, ctl_name, cf(9, bold=True, col="0F2A47"), lft()),
            (5, ctl_description, cf(9), lft()),
        ]
        for col, val, font_, align in cells:
            c = ws_t.cell(row=t_row, column=col, value=val)
            c.font = font_; c.alignment = align; c.border = bdr()
            c.fill = fill(bg)
        ws_t.row_dimensions[t_row].height = 60
        t_row += 1

ws_t.auto_filter.ref = f"A2:E{t_row - 1}"

# Hide the three matrix sheets — internal/reference only
_ws_lik.sheet_state = "hidden"
_ws_imp.sheet_state = "hidden"
ws_s.sheet_state = "hidden"

wb.save(OUTPUT_PATH)
if TAILORED:
    print(f"Tailored workbook saved → {OUTPUT_PATH}")
    print(f"  Industry: {INDUSTRY_NAMES[INDUSTRY_FILTER]} ({INDUSTRY_FILTER})")
    if NPAT:
        print(f"  NPAT: {_fmt_money(NPAT)} → impact scale calibrated")
else:
    print(f"Master library saved → {OUTPUT_PATH}")
print(f"Saved. Industry Profiles: 280 rows. Controls Library: ~{current_row} rows.")
print(f"Themes with controls: {len(CONTROLS)} of 40")
print(f"Themes with guidance: {len(GUIDANCE)} of 40")
