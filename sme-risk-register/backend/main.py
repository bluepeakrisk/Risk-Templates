"""
SME Risk Register — Backend API
Generates customised Excel risk registers based on user configuration.
"""
from fastapi import FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from typing import List, Optional
from io import BytesIO
import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

app = FastAPI(
    title="SME Risk Register API",
    description="Generate customised risk registers for small and medium enterprises",
    version="1.0.0",
)

# CORS — allow the frontend to call this API from anywhere
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],   # In production, lock this to your frontend domain
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


# ── Pydantic models for the request payload ────────────────────────────
class OrganisationConfig(BaseModel):
    name: str = "SME"
    industry: str = "Professional Services"
    size: str = "Medium (20-199 staff)"
    country: str = "Australia"
    cycle: str = "Quarterly"
    notes: Optional[str] = ""


class FinancialConfig(BaseModel):
    npat: float = 2000000
    currency: str = "AUD"
    fin_pcts: List[float] = [0.01, 0.03, 0.08, 0.15, 0.25]


class RAGThreshold(BaseModel):
    label: str
    min: int
    max: int


class RegisterConfig(BaseModel):
    organisation: OrganisationConfig
    financial: FinancialConfig
    focus_areas: List[str] = ["Cyber & Data", "Financial"]
    rag_thresholds: List[RAGThreshold] = [
        RAGThreshold(label="Green", min=1, max=5),
        RAGThreshold(label="Amber", min=6, max=14),
        RAGThreshold(label="Red", min=15, max=25),
    ]
    scoring_method: str = "max"  # max, avg, weighted


# ── Health check ──────────────────────────────────────────────────────
@app.get("/")
def root():
    return {"status": "ok", "service": "SME Risk Register API"}


@app.get("/health")
def health():
    return {"status": "healthy", "timestamp": datetime.datetime.utcnow().isoformat()}


# ── Main endpoint ──────────────────────────────────────────────────────
@app.post("/generate")
def generate_register(config: RegisterConfig):
    """Generate a customised Excel risk register and return as a download."""
    try:
        wb = build_workbook(config)
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)

        org_slug = config.organisation.name.replace(" ", "_").replace("/", "_")
        filename = f"Risk_Register_{org_slug}_{datetime.date.today().isoformat()}.xlsx"

        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Generation failed: {str(e)}")


# ── Workbook builder ───────────────────────────────────────────────────
def build_workbook(cfg: RegisterConfig) -> Workbook:
    """Construct the full risk register workbook from configuration."""
    wb = Workbook()

    # Style helpers
    C_NAVY = "1F3864"
    C_MBLUE = "2E5FA3"
    C_LBLUE = "D6E4F0"
    C_WHITE = "FFFFFF"
    C_LGREY = "F2F2F2"
    C_DGREY = "7F7F7F"
    C_GREEN = "27AE60"
    C_AMBER = "E67E22"
    C_RED = "C0392B"
    SCORE_COLORS = ["27AE60", "82C45E", "F0C040", "E67E22", "C0392B"]

    def bdr():
        s = Side(style="thin", color="BFBFBF")
        return Border(left=s, right=s, top=s, bottom=s)

    def fill(c): return PatternFill("solid", fgColor=c)
    def hf(sz=11, bold=True, col=C_WHITE): return Font(name="Arial", size=sz, bold=bold, color=col)
    def cf(sz=10, bold=False, col="000000"): return Font(name="Arial", size=sz, bold=bold, color=col)
    def ctr(): return Alignment(horizontal="center", vertical="center", wrap_text=True)
    def lft(): return Alignment(horizontal="left", vertical="center", wrap_text=True)

    # ── SHEET 1: COVER / INSTRUCTIONS ─────────────────────────────
    ws1 = wb.active
    ws1.title = "Instructions"
    ws1.sheet_view.showGridLines = False
    for col, w in [("A", 3), ("B", 28), ("C", 58), ("D", 20)]:
        ws1.column_dimensions[col].width = w

    ws1.merge_cells("B1:D2")
    c = ws1["B1"]
    c.value = f"Risk Register  |  {cfg.organisation.name}"
    c.font = Font(name="Arial", size=18, bold=True, color=C_WHITE)
    c.fill = fill(C_NAVY)
    c.alignment = ctr()
    ws1.row_dimensions[1].height = 22
    ws1.row_dimensions[2].height = 22

    ws1.merge_cells("B3:D3")
    c = ws1["B3"]
    c.value = f"{cfg.organisation.industry}  |  {cfg.organisation.size}  |  {cfg.organisation.country}"
    c.font = Font(name="Arial", size=11, italic=True, color=C_WHITE)
    c.fill = fill(C_MBLUE)
    c.alignment = ctr()
    ws1.row_dimensions[3].height = 18

    # Profile summary
    ws1.merge_cells("B5:D5")
    c = ws1["B5"]
    c.value = "ORGANISATION PROFILE"
    c.font = hf(sz=12, col=C_NAVY); c.fill = fill(C_LBLUE); c.alignment = lft()
    ws1.row_dimensions[5].height = 20

    profile_rows = [
        ("Organisation", cfg.organisation.name),
        ("Industry", cfg.organisation.industry),
        ("Size", cfg.organisation.size),
        ("Country", cfg.organisation.country),
        ("NPAT", f"{cfg.financial.currency} {cfg.financial.npat:,.0f}"),
        ("Catastrophic threshold", f"{cfg.financial.currency} {cfg.financial.npat * 0.25:,.0f} (25% NPAT)"),
        ("Review cycle", cfg.organisation.cycle),
        ("Focus areas", ", ".join(cfg.focus_areas)),
        ("Generated on", datetime.date.today().isoformat()),
    ]
    for i, (label, val) in enumerate(profile_rows, 6):
        bg = C_LGREY if i % 2 == 0 else C_WHITE
        c1 = ws1.cell(row=i, column=2, value=label)
        c1.font = cf(10, bold=True); c1.fill = fill(bg); c1.alignment = lft(); c1.border = bdr()
        c2 = ws1.cell(row=i, column=3, value=val)
        c2.font = cf(10); c2.fill = fill(bg); c2.alignment = lft(); c2.border = bdr()
        ws1.row_dimensions[i].height = 18

    # Sheets guide
    ws1.merge_cells("B17:D17")
    c = ws1["B17"]
    c.value = "SHEETS GUIDE"
    c.font = hf(sz=12, col=C_NAVY); c.fill = fill(C_LBLUE); c.alignment = lft()
    ws1.row_dimensions[17].height = 20

    sheets_info = [
        ("Instructions", "You are here. Read before starting."),
        ("Risk Scoring Setup", "Likelihood, 7 impact categories, RAG thresholds."),
        ("Risk Register", "Log all risks with auto-calculated scores."),
        ("Heat Map and Report", "Heat map and summary."),
        ("Treatment Tracker", "Track actions and owners."),
    ]
    for i, (sheet, purpose) in enumerate(sheets_info, 18):
        bg = C_LGREY if i % 2 == 0 else C_WHITE
        c1 = ws1.cell(row=i, column=2, value=sheet)
        c1.font = cf(10, bold=True); c1.fill = fill(bg); c1.alignment = lft(); c1.border = bdr()
        c2 = ws1.cell(row=i, column=3, value=purpose)
        c2.font = cf(10); c2.fill = fill(bg); c2.alignment = lft(); c2.border = bdr()
        ws1.row_dimensions[i].height = 18

    if cfg.organisation.notes:
        ws1.merge_cells("B25:D25")
        c = ws1["B25"]
        c.value = "NOTES"; c.font = hf(sz=12, col=C_NAVY); c.fill = fill(C_LBLUE); c.alignment = lft()
        ws1.row_dimensions[25].height = 20
        ws1.merge_cells("B26:D28")
        c = ws1["B26"]
        c.value = cfg.organisation.notes
        c.font = cf(10); c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws1.row_dimensions[26].height = 24

    # ── SHEET 2: SCORING SETUP ──────────────────────────────────
    ws2 = wb.create_sheet("Risk Scoring Setup")
    ws2.sheet_view.showGridLines = False
    for col, w in [("A", 7), ("B", 18), ("C", 40), ("D", 38)]:
        ws2.column_dimensions[col].width = w

    ws2.merge_cells("A1:D2")
    c = ws2["A1"]
    c.value = "Risk Scoring Setup"
    c.font = Font(name="Arial", size=16, bold=True, color=C_WHITE)
    c.fill = fill(C_NAVY); c.alignment = ctr()
    ws2.row_dimensions[1].height = 22; ws2.row_dimensions[2].height = 22

    # NPAT input row
    ws2.merge_cells("A4:D4")
    c = ws2["A4"]
    c.value = "NPAT INPUT (drives Financial impact thresholds)"
    c.font = hf(sz=11, col=C_NAVY); c.fill = fill("FFF2CC"); c.alignment = lft()
    ws2.row_dimensions[4].height = 22

    ws2.merge_cells("A5:B5")
    c = ws2["A5"]
    c.value = f"NPAT ({cfg.financial.currency})"
    c.font = hf(sz=10, col="7D4E00"); c.fill = fill("FFF2CC"); c.alignment = ctr(); c.border = bdr()

    npat_cell = ws2.cell(row=5, column=3, value=cfg.financial.npat)
    npat_cell.font = Font(name="Arial", size=11, bold=True, color="000080")
    npat_cell.fill = fill("EBF5FB"); npat_cell.alignment = ctr(); npat_cell.border = bdr()
    npat_cell.number_format = '#,##0'

    cat_cell = ws2.cell(row=5, column=4, value="=C5*0.25")
    cat_cell.font = Font(name="Arial", size=11, bold=True, color="7D0000")
    cat_cell.fill = fill("FADBD8"); cat_cell.alignment = ctr(); cat_cell.border = bdr()
    cat_cell.number_format = f'"{cfg.financial.currency} "#,##0'
    ws2.row_dimensions[5].height = 26

    # Section header helper
    def section(ws, row, text, span=4):
        ws.merge_cells(f"A{row}:{get_column_letter(span)}{row}")
        c = ws[f"A{row}"]
        c.value = text; c.font = hf(sz=11, col=C_NAVY); c.fill = fill(C_LBLUE); c.alignment = lft()
        ws.row_dimensions[row].height = 20

    def cat_table(ws, start_row, header, rows, finance=False):
        section(ws, start_row, header)
        # column headers
        col_hdrs = ["Score", "Rating", "Description", "Threshold" if finance else "Indicative measure"]
        for j, h in enumerate(col_hdrs, 1):
            c = ws.cell(row=start_row + 1, column=j, value=h)
            c.font = hf(sz=9); c.fill = fill(C_MBLUE); c.alignment = ctr(); c.border = bdr()
        ws.row_dimensions[start_row + 1].height = 28
        for i, (sc, rat, desc, metric) in enumerate(rows):
            r = start_row + 2 + i
            bg = C_LGREY if r % 2 == 0 else C_WHITE
            sc_bg = SCORE_COLORS[sc - 1]
            c = ws.cell(row=r, column=1, value=sc); c.font = hf(sz=10); c.fill = fill(sc_bg); c.alignment = ctr(); c.border = bdr()
            c = ws.cell(row=r, column=2, value=rat); c.font = hf(sz=9, bold=False); c.fill = fill(sc_bg); c.alignment = ctr(); c.border = bdr()
            c = ws.cell(row=r, column=3, value=desc); c.font = cf(9); c.fill = fill(bg); c.alignment = lft(); c.border = bdr()
            if finance:
                c = ws.cell(row=r, column=4, value=f"=C5*{cfg.financial.fin_pcts[i]}")
                c.number_format = f'"Up to {cfg.financial.currency} "#,##0'
            else:
                c = ws.cell(row=r, column=4, value=metric)
            c.font = cf(9, col="555555"); c.fill = fill(bg); c.alignment = lft(); c.border = bdr()
            ws.row_dimensions[r].height = 26

    # Financial
    fin_rows = [
        (1, "Negligible", "Immaterial loss; absorbed in normal operations.", ""),
        (2, "Minor", "Small loss; manageable within budget.", ""),
        (3, "Moderate", "Significant loss; management action required.", ""),
        (4, "Major", "Serious loss; board attention required.", ""),
        (5, "Catastrophic", "Threatens solvency or business continuity.", ""),
    ]
    cat_table(ws2, 7, "1. FINANCIAL IMPACT (auto-calculated from NPAT)", fin_rows, finance=True)

    # Customers
    cust_rows = [
        (1, "Negligible", "Single complaint; resolved immediately.", "<0.1% customers affected"),
        (2, "Minor", "Small number dissatisfied; minimal churn.", "<1% customers affected"),
        (3, "Moderate", "Noticeable dissatisfaction; measurable churn.", "1–5% customers affected"),
        (4, "Major", "Significant customer loss; key accounts threatened.", "5–20% at risk"),
        (5, "Catastrophic", "Mass exodus; brand irreparably damaged.", ">20% lost"),
    ]
    cat_table(ws2, 15, "2. CUSTOMER IMPACT", cust_rows)

    # Regulatory
    reg_rows = [
        (1, "Negligible", "Minor procedural breach; self-corrected.", "Verbal warning only"),
        (2, "Minor", "Low-level non-compliance; remediation required.", "Written notice; minor fine"),
        (3, "Moderate", "Formal sanction; public disclosure likely.", "Moderate fine; remediation order"),
        (4, "Major", "Serious breach; large fine; licence suspension.", "Significant fine; suspension"),
        (5, "Catastrophic", "Licence revocation; criminal prosecution.", "Criminal liability; closure"),
    ]
    cat_table(ws2, 23, "3. REGULATORY IMPACT", reg_rows)

    # Reputational
    rep_rows = [
        (1, "Negligible", "Isolated negative comment; no media.", "Internal only"),
        (2, "Minor", "Limited negative coverage; quickly rebutted.", "Local media; brief"),
        (3, "Moderate", "Sustained media; stakeholder concern.", "Regional media; formal response"),
        (4, "Major", "National coverage; significant backlash.", "National press; peak body scrutiny"),
        (5, "Catastrophic", "Severe brand destruction; loss of trust.", "International coverage; existential"),
    ]
    cat_table(ws2, 31, "4. REPUTATIONAL IMPACT", rep_rows)

    # People
    ppl_rows = [
        (1, "Negligible", "Minor inconvenience; no injuries.", "Near miss only"),
        (2, "Minor", "Minor injury; first aid; brief absence.", "First aid; <1 day lost"),
        (3, "Moderate", "Moderate injury; medical treatment; reportable.", "1–5 days lost; notifiable"),
        (4, "Major", "Serious injury; hospitalisation; notifiable.", ">5 days lost; regulator notified"),
        (5, "Catastrophic", "Fatality or permanent disability.", "Fatality / permanent incapacity"),
    ]
    cat_table(ws2, 39, "5. PEOPLE IMPACT (Health, Safety, Wellbeing)", ppl_rows)

    # Social
    soc_rows = [
        (1, "Negligible", "No community concern raised.", "No external complaints"),
        (2, "Minor", "Limited concern; informal contact.", "Isolated complaint"),
        (3, "Moderate", "Organised concern; public meetings.", "Petition; local government"),
        (4, "Major", "Significant harm; public opposition.", "Government intervention"),
        (5, "Catastrophic", "Severe lasting harm; loss of social licence.", "Forced exit"),
    ]
    cat_table(ws2, 47, "6. SOCIAL IMPACT (External Stakeholders)", soc_rows)

    # Environmental
    env_rows = [
        (1, "Negligible", "Trivial, contained effect.", "Self-contained"),
        (2, "Minor", "Minor, localised; quickly remediated.", "Rapidly remediated"),
        (3, "Moderate", "Measurable harm; regulator notified.", "Localised contamination"),
        (4, "Major", "Significant damage; extended remediation.", "Protected ecosystem affected"),
        (5, "Catastrophic", "Irreversible destruction; prosecution.", "Criminal prosecution; shutdown"),
    ]
    cat_table(ws2, 55, "7. ENVIRONMENTAL IMPACT", env_rows)

    # Likelihood
    lh_rows = [
        (1, "Rare", "May only occur in exceptional circumstances.", "<1 in 10 years"),
        (2, "Unlikely", "Could occur but not expected.", "1 in 5–10 years"),
        (3, "Possible", "Might occur at some point.", "1 in 2–5 years"),
        (4, "Likely", "Will probably occur.", "Once a year"),
        (5, "Almost Certain", "Expected to occur frequently.", "Multiple times per year"),
    ]
    cat_table(ws2, 63, "LIKELIHOOD SCALE", lh_rows)

    # RAG thresholds
    section(ws2, 71, "RAG THRESHOLDS")
    rag_hdrs = ["RAG", "Score range", "Meaning", "Required action"]
    for j, h in enumerate(rag_hdrs, 1):
        c = ws2.cell(row=72, column=j, value=h); c.font = hf(sz=9); c.fill = fill(C_MBLUE); c.alignment = ctr(); c.border = bdr()
    rag_meta = [
        (cfg.rag_thresholds[0], C_GREEN, "Low — within appetite", "Monitor; review annually"),
        (cfg.rag_thresholds[1], C_AMBER, "Medium — approaching threshold", "Treat; review quarterly"),
        (cfg.rag_thresholds[2], C_RED,   "High — outside appetite", "Immediate action; escalate"),
    ]
    for i, (rag, color, meaning, action) in enumerate(rag_meta):
        r = 73 + i
        c = ws2.cell(row=r, column=1, value=rag.label.upper()); c.font = hf(sz=10); c.fill = fill(color); c.alignment = ctr(); c.border = bdr()
        c = ws2.cell(row=r, column=2, value=f"{rag.min} – {rag.max}"); c.font = hf(sz=10); c.fill = fill(color); c.alignment = ctr(); c.border = bdr()
        c = ws2.cell(row=r, column=3, value=meaning); c.font = cf(9); c.alignment = lft(); c.border = bdr()
        c = ws2.cell(row=r, column=4, value=action); c.font = cf(9); c.alignment = lft(); c.border = bdr()
        ws2.row_dimensions[r].height = 24

    # ── SHEET 3: RISK REGISTER ──────────────────────────────────
    ws3 = wb.create_sheet("Risk Register")
    ws3.sheet_view.showGridLines = False
    ws3.freeze_panes = "A4"

    col_widths3 = [
        ("A", 7), ("B", 20), ("C", 28), ("D", 38), ("E", 22), ("F", 22),
        ("G", 13), ("H", 13), ("I", 13), ("J", 14),
        ("K", 26), ("L", 16),
        ("M", 13), ("N", 13), ("O", 13), ("P", 14),
        ("Q", 20), ("R", 16), ("S", 18), ("T", 22),
    ]
    for col, w in col_widths3:
        ws3.column_dimensions[col].width = w

    ws3.merge_cells("A1:T1")
    c = ws3["A1"]
    c.value = f"Risk Register  |  {cfg.organisation.name}  |  Generated {datetime.date.today().isoformat()}"
    c.font = Font(name="Arial", size=14, bold=True, color=C_WHITE)
    c.fill = fill(C_NAVY); c.alignment = ctr()
    ws3.row_dimensions[1].height = 26

    groups = [
        ("A2:A2", "", ""),
        ("B2:F2", "RISK IDENTIFICATION", C_MBLUE),
        ("G2:J2", "INHERENT RISK", C_RED),
        ("K2:L2", "EXISTING CONTROLS", "148F77"),
        ("M2:P2", "RESIDUAL RISK", C_AMBER),
        ("Q2:T2", "OWNERSHIP & TREATMENT", C_NAVY),
    ]
    for merge_rng, label, color in groups:
        ws3.merge_cells(merge_rng)
        start = merge_rng.split(":")[0]
        c = ws3[start]; c.value = label
        c.font = hf(sz=10); c.fill = fill(color if color else C_NAVY); c.alignment = ctr()
    ws3.row_dimensions[2].height = 20

    headers3 = [
        "Risk ID", "Category", "Risk Title", "Risk Description",
        "Potential Cause", "Potential Consequence",
        "Likelihood\n(1-5)", "Impact\n(1-5)", "Inherent\nScore", "Inherent\nRAG",
        "Existing Controls", "Control\nEffectiveness",
        "Likelihood\n(1-5)", "Impact\n(1-5)", "Residual\nScore", "Residual\nRAG",
        "Risk Owner", "Next Review", "Treatment\nStrategy", "Comments"
    ]
    hdr_bgs = [C_NAVY] + [C_MBLUE] * 5 + [C_RED] * 4 + ["148F77"] * 2 + [C_AMBER] * 4 + [C_NAVY] * 4
    for j, (hdr, bg) in enumerate(zip(headers3, hdr_bgs), 1):
        c = ws3.cell(row=3, column=j, value=hdr)
        c.font = hf(sz=9); c.fill = fill(bg); c.alignment = ctr(); c.border = bdr()
    ws3.row_dimensions[3].height = 36

    # Dropdowns
    cat_list = "Strategic,Financial,Operational,People and HR,Technology and Cyber,Legal and Regulatory,Reputational,Environmental and Social,Health and Safety,Project and Change,Third Party and Supply,Customer and Market,Geopolitical and Economic"
    dv_cat = DataValidation(type="list", formula1=f'"{cat_list}"', allow_blank=True)
    dv_cat.sqref = "B4:B103"; ws3.add_data_validation(dv_cat)

    dv_li = DataValidation(type="list", formula1='"1,2,3,4,5"', allow_blank=True)
    dv_li.sqref = "G4:H103 M4:N103"; ws3.add_data_validation(dv_li)

    dv_ctrl = DataValidation(type="list", formula1='"Strong,Adequate,Weak,None"', allow_blank=True)
    dv_ctrl.sqref = "L4:L103"; ws3.add_data_validation(dv_ctrl)

    dv_trt = DataValidation(type="list", formula1='"Avoid,Reduce,Transfer,Accept"', allow_blank=True)
    dv_trt.sqref = "S4:S103"; ws3.add_data_validation(dv_trt)

    # Generate tailored risks based on focus areas
    sample_risks = generate_tailored_risks(cfg)

    CENTER_COLS = {1, 7, 8, 9, 10, 12, 13, 14, 15, 16}
    rag_min = cfg.rag_thresholds[2].min  # Red threshold
    amber_min = cfg.rag_thresholds[1].min

    for row_idx, risk in enumerate(sample_risks, 4):
        bg = C_LGREY if row_idx % 2 == 0 else C_WHITE
        for col_idx, val in enumerate(risk, 1):
            c = ws3.cell(row=row_idx, column=col_idx)
            c.font = cf(9); c.fill = fill(bg); c.border = bdr()
            c.alignment = ctr() if col_idx in CENTER_COLS else lft()
            if col_idx == 9:
                c.value = f'=IF(AND(G{row_idx}<>"",H{row_idx}<>""),G{row_idx}*H{row_idx},"")'
            elif col_idx == 10:
                c.value = f'=IF(I{row_idx}="","",IF(I{row_idx}>={rag_min},"RED",IF(I{row_idx}>={amber_min},"AMBER","GREEN")))'
            elif col_idx == 15:
                c.value = f'=IF(AND(M{row_idx}<>"",N{row_idx}<>""),M{row_idx}*N{row_idx},"")'
            elif col_idx == 16:
                c.value = f'=IF(O{row_idx}="","",IF(O{row_idx}>={rag_min},"RED",IF(O{row_idx}>={amber_min},"AMBER","GREEN")))'
            else:
                c.value = val
        ws3.row_dimensions[row_idx].height = 40

    # Empty rows with formulas
    for row_idx in range(4 + len(sample_risks), 104):
        bg = C_LGREY if row_idx % 2 == 0 else C_WHITE
        for col_idx in range(1, 21):
            c = ws3.cell(row=row_idx, column=col_idx)
            c.fill = fill(bg); c.border = bdr(); c.font = cf(9)
            c.alignment = ctr() if col_idx in CENTER_COLS else lft()
            if col_idx == 9:
                c.value = f'=IF(AND(G{row_idx}<>"",H{row_idx}<>""),G{row_idx}*H{row_idx},"")'
            elif col_idx == 10:
                c.value = f'=IF(I{row_idx}="","",IF(I{row_idx}>={rag_min},"RED",IF(I{row_idx}>={amber_min},"AMBER","GREEN")))'
            elif col_idx == 15:
                c.value = f'=IF(AND(M{row_idx}<>"",N{row_idx}<>""),M{row_idx}*N{row_idx},"")'
            elif col_idx == 16:
                c.value = f'=IF(O{row_idx}="","",IF(O{row_idx}>={rag_min},"RED",IF(O{row_idx}>={amber_min},"AMBER","GREEN")))'
        ws3.row_dimensions[row_idx].height = 30

    # ── SHEET 4: HEAT MAP & REPORT ────────────────────────────
    ws4 = wb.create_sheet("Heat Map and Report")
    ws4.sheet_view.showGridLines = False
    for col in list("ABCDEFGHIJKLMNO"):
        ws4.column_dimensions[col].width = 14

    ws4.merge_cells("A1:O2")
    c = ws4["A1"]
    c.value = "Risk Heat Map and Executive Summary"
    c.font = Font(name="Arial", size=16, bold=True, color=C_WHITE)
    c.fill = fill(C_NAVY); c.alignment = ctr()
    ws4.row_dimensions[1].height = 22; ws4.row_dimensions[2].height = 22

    # Heat map grid
    heat_colors = {
        (5, 1): "F39C12", (5, 2): "E74C3C", (5, 3): "E74C3C", (5, 4): "C0392B", (5, 5): "C0392B",
        (4, 1): "F39C12", (4, 2): "F39C12", (4, 3): "E74C3C", (4, 4): "E74C3C", (4, 5): "C0392B",
        (3, 1): "27AE60", (3, 2): "F39C12", (3, 3): "F39C12", (3, 4): "E74C3C", (3, 5): "E74C3C",
        (2, 1): "27AE60", (2, 2): "27AE60", (2, 3): "F39C12", (2, 4): "F39C12", (2, 5): "E74C3C",
        (1, 1): "27AE60", (1, 2): "27AE60", (1, 3): "27AE60", (1, 4): "27AE60", (1, 5): "F39C12",
    }

    ws4.merge_cells("A4:F4")
    c = ws4["A4"]
    c.value = "RESIDUAL RISK HEAT MAP"
    c.font = hf(sz=12, col=C_NAVY); c.fill = fill(C_LBLUE); c.alignment = lft()
    ws4.row_dimensions[4].height = 20

    ws4.cell(row=5, column=1, value="IMPACT ↓ / LIKELIHOOD →").font = hf(sz=9)
    c = ws4.cell(row=5, column=1); c.fill = fill(C_NAVY); c.alignment = ctr(); c.border = bdr()
    likelihood_lbls = ["1 - Rare", "2 - Unlikely", "3 - Possible", "4 - Likely", "5 - Almost Certain"]
    for li_idx, lbl in enumerate(likelihood_lbls, 2):
        c = ws4.cell(row=5, column=li_idx, value=lbl); c.font = hf(sz=9); c.fill = fill(C_MBLUE); c.alignment = ctr(); c.border = bdr()
    ws4.row_dimensions[5].height = 28

    impact_lbls = ["5 - Catastrophic", "4 - Major", "3 - Moderate", "2 - Minor", "1 - Negligible"]
    for r_off, (impact_val, imp_lbl) in enumerate(zip([5, 4, 3, 2, 1], impact_lbls), 6):
        c = ws4.cell(row=r_off, column=1, value=imp_lbl); c.font = hf(sz=9); c.fill = fill(C_MBLUE); c.alignment = ctr(); c.border = bdr()
        ws4.row_dimensions[r_off].height = 36
        for lh_col, lh_val in enumerate([1, 2, 3, 4, 5], 2):
            score = impact_val * lh_val
            color = heat_colors[(impact_val, lh_val)]
            cell = ws4.cell(row=r_off, column=lh_col, value=score)
            cell.fill = fill(color)
            cell.font = Font(name="Arial", size=12, bold=True, color=C_WHITE)
            cell.alignment = ctr(); cell.border = bdr()

    # Summary panel
    ws4.merge_cells("H4:O4")
    c = ws4["H4"]
    c.value = "RISK SUMMARY (auto-calculated from Risk Register)"
    c.font = hf(sz=11, col=C_NAVY); c.fill = fill(C_LBLUE); c.alignment = lft()

    summary_items = [
        ("Total Risks Logged", "=COUNTA('Risk Register'!C4:C103)"),
        ("RED Risks", "=COUNTIF('Risk Register'!P4:P103,\"RED\")"),
        ("AMBER Risks", "=COUNTIF('Risk Register'!P4:P103,\"AMBER\")"),
        ("GREEN Risks", "=COUNTIF('Risk Register'!P4:P103,\"GREEN\")"),
        ("Average Residual Score", "=IFERROR(ROUND(AVERAGE('Risk Register'!O4:O103),1),\"N/A\")"),
    ]
    bg_map = {"RED Risks": "FADBD8", "AMBER Risks": "FDEBD0", "GREEN Risks": "D5F5E3"}
    for i, (label, formula) in enumerate(summary_items, 5):
        ws4.merge_cells(f"H{i}:K{i}")
        c1 = ws4[f"H{i}"]; c1.value = label; c1.font = cf(10, bold=True); c1.fill = fill(C_LBLUE); c1.alignment = lft(); c1.border = bdr()
        ws4.merge_cells(f"L{i}:M{i}")
        c2 = ws4[f"L{i}"]; c2.value = formula; c2.font = cf(11, bold=True); c2.alignment = ctr(); c2.border = bdr()
        c2.fill = fill(bg_map.get(label, C_LGREY))
        ws4.row_dimensions[i].height = 22

    # ── SHEET 5: TREATMENT TRACKER ───────────────────────────
    ws5 = wb.create_sheet("Treatment Tracker")
    ws5.sheet_view.showGridLines = False
    ws5.freeze_panes = "A3"
    for col, w in [("A", 8), ("B", 14), ("C", 26), ("D", 40), ("E", 18), ("F", 18), ("G", 18), ("H", 18), ("I", 16), ("J", 14), ("K", 26)]:
        ws5.column_dimensions[col].width = w

    ws5.merge_cells("A1:K1")
    c = ws5["A1"]
    c.value = "Treatment Tracker"
    c.font = Font(name="Arial", size=14, bold=True, color=C_WHITE)
    c.fill = fill(C_NAVY); c.alignment = ctr()
    ws5.row_dimensions[1].height = 26

    trt_hdrs = ["Action Ref", "Risk ID", "Risk Title", "Action Description", "Strategy", "Owner", "Target Date", "Actual Date", "Status", "% Complete", "Notes"]
    for j, hdr in enumerate(trt_hdrs, 1):
        c = ws5.cell(row=2, column=j, value=hdr); c.font = hf(sz=9); c.fill = fill(C_MBLUE); c.alignment = ctr(); c.border = bdr()
    ws5.row_dimensions[2].height = 32

    dv_status = DataValidation(type="list", formula1='"Not Started,In Progress,Complete,Overdue,On Hold"', allow_blank=True)
    dv_status.sqref = "I3:I103"; ws5.add_data_validation(dv_status)
    dv_pct = DataValidation(type="list", formula1='"0%,25%,50%,75%,100%"', allow_blank=True)
    dv_pct.sqref = "J3:J103"; ws5.add_data_validation(dv_pct)
    dv_strat = DataValidation(type="list", formula1='"Avoid,Reduce,Transfer,Accept"', allow_blank=True)
    dv_strat.sqref = "E3:E103"; ws5.add_data_validation(dv_strat)

    for r in range(3, 103):
        bg = C_LGREY if r % 2 == 0 else C_WHITE
        for col_idx in range(1, 12):
            c = ws5.cell(row=r, column=col_idx); c.fill = fill(bg); c.border = bdr(); c.font = cf(9)
            c.alignment = ctr() if col_idx in [1, 2, 5, 9, 10] else lft()
        ws5.row_dimensions[r].height = 26

    return wb


def generate_tailored_risks(cfg: RegisterConfig) -> list:
    """Generate a starter set of risks based on focus areas."""
    risks = []
    focus = set(cfg.focus_areas)
    rid = 1

    cyber_risks = [
        ("Technology and Cyber", "Phishing & Business Email Compromise",
         "Staff deceived into transferring funds or revealing credentials via fraudulent emails.",
         "Sophisticated phishing; weak email authentication; low staff awareness.",
         "Direct financial loss; data exposure; regulatory notification.",
         4, 5, "Email filtering; annual awareness training", "Weak", 3, 4,
         "IT Manager", "2025-09-30", "Reduce", "Deploy MFA, DMARC, phishing simulation"),
        ("Technology and Cyber", "Ransomware Attack",
         "Malicious software encrypts systems and data, with attackers demanding payment.",
         "Unpatched systems; phishing; inadequate endpoint protection.",
         "Operational shutdown; data loss; regulatory fines.",
         4, 5, "Daily backups; antivirus; firewall", "Weak", 3, 5,
         "IT Manager", "2025-09-30", "Reduce", "Implement EDR; test DR plan"),
        ("Technology and Cyber", "Data Breach",
         "Unauthorised access to confidential client or company data.",
         "Insider threat; stolen credentials; misconfigured cloud.",
         "Notifiable breach; regulatory fines; lawsuits; loss of contracts.",
         3, 5, "RBAC; NDAs; basic encryption", "Adequate", 2, 4,
         "Privacy Officer", "2025-09-30", "Reduce", "Conduct data mapping; breach response plan"),
    ]

    financial_risks = [
        ("Financial", "Revenue Concentration — Key Customer Loss",
         "Loss of one or two major customers representing a disproportionate share of revenue.",
         "Customer insolvency; competitor poaching; service quality failure.",
         "Revenue shortfall; cash flow crisis; staff redundancies.",
         3, 4, "Diversified base (partial); strong relationships", "Adequate", 2, 4,
         "Managing Director", "2025-09-30", "Reduce", "Monitor top-5 concentration; target <20% per customer"),
        ("Financial", "Cash Flow Shortfall",
         "Insufficient liquidity to meet short-term obligations.",
         "Delayed receivables; unexpected expenses.",
         "Inability to pay suppliers/staff; disruption.",
         3, 4, "Monthly forecast; credit facility", "Adequate", 2, 3,
         "CFO", "2025-09-30", "Reduce", "Review credit limit annually"),
        ("Financial", "Fraud or Misappropriation",
         "Internal fraud by staff or management involving misuse of funds.",
         "Inadequate segregation of duties; trust-based culture.",
         "Direct loss; regulatory action; reputational destruction.",
         2, 5, "Dual-authorisation on payments; annual audit", "Adequate", 1, 4,
         "CFO", "2025-09-30", "Reduce", "Formalise expense matrix; surprise audits"),
    ]

    operational_risks = [
        ("Operational", "Supply Chain Disruption",
         "Failure or delay of key suppliers impacting delivery.",
         "Single-source suppliers; geopolitical events.",
         "Revenue loss; customer dissatisfaction.",
         3, 3, "Preferred supplier agreements", "Weak", 3, 3,
         "Operations Manager", "2025-09-30", "Reduce", "Identify alternatives for top 5 inputs"),
    ]

    people_risks = [
        ("People and HR", "Loss of Key People",
         "Departure of senior staff with critical knowledge or relationships.",
         "Competitive market; below-market remuneration; culture issues.",
         "Capability loss; client relationship risk; team morale impact.",
         4, 4, "Employment contracts; some succession", "Weak", 3, 4,
         "Managing Director", "2025-09-30", "Reduce", "Implement retention program"),
        ("Health and Safety", "Workplace Injury",
         "Employee or visitor injured on premises or during operations.",
         "Unsafe practices; inadequate training.",
         "Worker compensation; legal liability.",
         2, 3, "OH&S policy; basic training", "Adequate", 1, 3,
         "Operations Manager", "2025-09-30", "Reduce", "Complete formal WHS audit"),
    ]

    regulatory_risks = [
        ("Legal and Regulatory", "Regulatory Non-Compliance",
         "Failure to comply with applicable laws or regulations.",
         "Rapid regulatory change; lack of legal expertise.",
         "Fines; sanctions; licence issues.",
         3, 4, "Legal counsel; compliance checklist", "Adequate", 2, 3,
         "Operations Manager", "2025-12-31", "Reduce", "Subscribe to regulatory updates"),
    ]

    reputational_risks = [
        ("Reputational", "Negative PR or Customer Review",
         "Public criticism damaging firm reputation via media or online platforms.",
         "Service failure; billing dispute; misconduct.",
         "Loss of prospects; talent attraction issues.",
         3, 3, "Customer satisfaction; complaints handling", "Weak", 2, 3,
         "Marketing Manager", "2025-12-31", "Reduce", "Implement post-engagement NPS survey"),
    ]

    environmental_risks = [
        ("Environmental and Social", "Environmental Compliance Failure",
         "Failure to meet environmental obligations causing harm and breach.",
         "Inadequate controls; staff awareness gaps.",
         "Regulatory action; remediation costs; reputational damage.",
         2, 3, "Basic environmental policy", "Weak", 2, 3,
         "Operations Manager", "2025-12-31", "Reduce", "Conduct environmental risk assessment"),
    ]

    # Add risks based on focus areas
    if "Cyber & Data" in focus:
        risks.extend(cyber_risks)
    if "Financial" in focus:
        risks.extend(financial_risks)
    if "Operational" in focus:
        risks.extend(operational_risks)
    if "People & Safety" in focus:
        risks.extend(people_risks)
    if "Regulatory" in focus:
        risks.extend(regulatory_risks)
    if "Reputational" in focus:
        risks.extend(reputational_risks)
    if "Environmental" in focus:
        risks.extend(environmental_risks)

    # If no focus areas, give a balanced default set
    if not risks:
        risks = cyber_risks[:1] + financial_risks[:1] + operational_risks + people_risks + regulatory_risks + reputational_risks

    # Cap at 15 risks, add IDs and convert to tuples with formula placeholders
    final = []
    for i, r in enumerate(risks[:15], 1):
        rid = f"R-{i:03d}"
        # category, title, desc, cause, conseq, inh_l, inh_i, controls, ctrl_eff, res_l, res_i, owner, review, strat, comments
        cat, title, desc, cause, conseq, inh_l, inh_i, controls, ctrl_eff, res_l, res_i, owner, review, strat, comments = r
        # Reorganise to match column order: id, cat, title, desc, cause, conseq, inh_l, inh_i, "", "", controls, ctrl_eff, res_l, res_i, "", "", owner, review, strat, comments
        final.append((rid, cat, title, desc, cause, conseq, inh_l, inh_i, "", "", controls, ctrl_eff, res_l, res_i, "", "", owner, review, strat, comments))

    return final


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
